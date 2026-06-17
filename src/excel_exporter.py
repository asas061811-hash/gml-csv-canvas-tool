"""
excel_exporter.py - 以 openpyxl 輸出五份 Excel：
  1. 解析後點位表.xlsx
  2. 異常清單.xlsx
  3. 重複量測比對表.xlsx
  4. 人工判讀紀錄範本.xlsx  (v2)
  5. 修正紀錄範本.xlsx       (v2)

  另提供 _to_bytes / export_*_df_bytes 系列函式供 Streamlit 使用。
"""

import io as _io
import os
import math
import tempfile
import pandas as pd
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ---- 樣式常數 ----
HEADER_FILL_BLUE   = PatternFill('solid', fgColor='1F4E79')
HEADER_FILL_RED    = PatternFill('solid', fgColor='922B21')
HEADER_FILL_ORANGE = PatternFill('solid', fgColor='7E5109')
HEADER_FILL_GREEN  = PatternFill('solid', fgColor='1A5632')   # v2 人工判讀
HEADER_FILL_PURPLE = PatternFill('solid', fgColor='4A235A')   # v2 修正紀錄
EDITABLE_COL_FILL  = PatternFill('solid', fgColor='EAF4FB')   # 可填寫欄位底色
ANOMALY_ROW_FILL   = PatternFill('solid', fgColor='FEF9E7')   # 預標異常列底色（淡黃）
HEADER_FONT = Font(color='FFFFFF', bold=True, name='微軟正黑體', size=10)
BODY_FONT   = Font(name='微軟正黑體', size=10)
BODY_FONT_RED = Font(name='微軟正黑體', size=10, color='C0392B')
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=False)
WRAP_ALIGN   = Alignment(horizontal='left', vertical='center', wrap_text=True)
LEFT_ALIGN   = Alignment(horizontal='left', vertical='center')

THIN_SIDE  = Side(style='thin', color='BBBBBB')
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

ROW_FILL_ALT  = PatternFill('solid', fgColor='EBF5FB')  # 淡藍（隔行）
ROW_FILL_ANOM = PatternFill('solid', fgColor='FDECEA')  # 淡紅（異常列）


def _apply_header(ws, headers, fill):
    """寫入標題列並套用樣式。"""
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def _auto_col_width(ws, min_w=8, max_w=40):
    """自動調整欄寬。"""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        adjusted = min(max_w, max(min_w, max_len + 2))
        ws.column_dimensions[col_letter].width = adjusted


def _write_rows(ws, rows_data, start_row=2, alt_fill=True, highlight_col=None, highlight_vals=None):
    """寫入資料列並套用交替底色；highlight_col 欄位值在 highlight_vals 內的列顯示紅字。"""
    for r_idx, row in enumerate(rows_data, start=start_row):
        is_alt = (r_idx % 2 == 0)
        is_highlight = False
        if highlight_col is not None and highlight_vals is not None:
            val = row[list(row.keys())[highlight_col]] if isinstance(row, dict) else None
            is_highlight = str(val) in highlight_vals if val is not None else False

        for c_idx, (key, val) in enumerate(row.items(), start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=_fmt(val))
            cell.border = THIN_BORDER
            cell.alignment = LEFT_ALIGN

            if is_highlight:
                cell.font = BODY_FONT_RED
                cell.fill = ROW_FILL_ANOM
            elif alt_fill and is_alt:
                cell.font = BODY_FONT
                cell.fill = ROW_FILL_ALT
            else:
                cell.font = BODY_FONT


def _fmt(val):
    """None / NaN → 空字串，其餘保留原值（數字格式不變）。"""
    import math
    if val is None:
        return ''
    if isinstance(val, float) and math.isnan(val):
        return ''
    if isinstance(val, (int, float)):
        return val
    return str(val)


# ============================================================
# 1. 解析後點位表
# ============================================================
PARSED_HEADERS = [
    '原始識別碼', '管線識別碼', '點號',
    '類別碼', 'X(E)', 'Y(N)', 'Z(H)',
    '測量日期', '來源CSV', '原始列號',
]


def export_parsed_points(df, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '解析後點位表'
    ws.freeze_panes = 'A2'

    _apply_header(ws, PARSED_HEADERS, HEADER_FILL_BLUE)

    rows_data = []
    for _, row in df.iterrows():
        rows_data.append({
            '原始識別碼': row.get('raw_id'),
            '管線識別碼': row.get('pipeline_id'),
            '點號':      row.get('point_no'),
            '類別碼':    row.get('raw_category'),
            'X(E)':     row.get('raw_x'),
            'Y(N)':     row.get('raw_y'),
            'Z(H)':     row.get('raw_z'),
            '測量日期':  row.get('raw_date'),
            '來源CSV':   row.get('_source_file'),
            '原始列號':  row.get('_original_row'),
        })

    _write_rows(ws, rows_data)
    _auto_col_width(ws)
    ws.row_dimensions[1].height = 20

    wb.save(output_path)
    print(f'  已輸出: {output_path}（{len(rows_data)} 筆）')


# ============================================================
# 2. 異常清單
# ============================================================
ANOMALY_HEADERS = [
    '異常類型', '異常說明',
    '原始識別碼', '管線識別碼', '點號',
    'X', 'Y', 'Z', '測量日期',
    '來源CSV', '原始列號', '建議處理方式',
]

# 需要用紅字標示的嚴重異常類型
HIGH_SEVERITY = {
    '座標空值', 'X 座標格式異常', 'Y 座標格式異常',
    '重複點位座標差異過大',
}


def export_anomalies(anomaly_list, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '異常清單'
    ws.freeze_panes = 'A2'

    _apply_header(ws, ANOMALY_HEADERS, HEADER_FILL_RED)

    for r_idx, item in enumerate(anomaly_list, start=2):
        is_high = item.get('異常類型') in HIGH_SEVERITY
        is_alt = (r_idx % 2 == 0)

        values = [
            item.get('異常類型'),
            item.get('異常說明'),
            item.get('原始識別碼'),
            item.get('管線識別碼'),
            item.get('點號'),
            item.get('X'),
            item.get('Y'),
            item.get('Z'),
            item.get('測量日期'),
            item.get('來源CSV'),
            item.get('原始列號'),
            item.get('建議處理方式'),
        ]
        for c_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=_fmt(val))
            cell.border = THIN_BORDER
            if is_high:
                cell.font = BODY_FONT_RED
                cell.fill = ROW_FILL_ANOM
                cell.alignment = LEFT_ALIGN
            elif is_alt:
                cell.font = BODY_FONT
                cell.fill = ROW_FILL_ALT
                cell.alignment = LEFT_ALIGN
            else:
                cell.font = BODY_FONT
                cell.alignment = LEFT_ALIGN

        # 「建議處理方式」欄位自動換行
        ws.cell(row=r_idx, column=12).alignment = WRAP_ALIGN

    _auto_col_width(ws)
    # 說明欄寬度固定較寬
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['L'].width = 50
    ws.row_dimensions[1].height = 20

    wb.save(output_path)
    print(f'  已輸出: {output_path}（{len(anomaly_list)} 筆異常）')


# ============================================================
# 3. 重複量測比對表
# ============================================================
DUPLICATE_HEADERS = [
    '來源CSV', '原始列號',
    '原始識別碼', '管線識別碼', '點號',
    '測量日期', 'X', 'Y', 'Z',
    '比對基準來源CSV', '比對基準原始列號',
    '與比對點XY差距(m)', '與比對點Z差距(m)',
    '判定結果', '建議處理方式',
]

VERDICT_COLORS = {
    '近似重複，待確認':         PatternFill('solid', fgColor='FEF9E7'),
    '重複點位座標差異過大，待確認': PatternFill('solid', fgColor='FDECEA'),
    '無法比對（座標缺漏）':      PatternFill('solid', fgColor='F2F3F4'),
}


def export_duplicates(duplicate_list, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '重複量測比對表'
    ws.freeze_panes = 'A2'

    _apply_header(ws, DUPLICATE_HEADERS, HEADER_FILL_ORANGE)

    for r_idx, item in enumerate(duplicate_list, start=2):
        verdict = str(item.get('判定結果', ''))
        row_fill = VERDICT_COLORS.get(verdict)

        values = [
            item.get('來源CSV'),
            item.get('原始列號'),
            item.get('原始識別碼'),
            item.get('管線識別碼'),
            item.get('點號'),
            item.get('測量日期'),
            item.get('X'),
            item.get('Y'),
            item.get('Z'),
            item.get('比對基準來源CSV'),
            item.get('比對基準原始列號'),
            item.get('與比對點XY差距(m)'),
            item.get('與比對點Z差距(m)'),
            verdict,
            item.get('建議處理方式'),
        ]
        for c_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=_fmt(val))
            cell.border = THIN_BORDER
            cell.font = BODY_FONT
            cell.alignment = LEFT_ALIGN
            if row_fill:
                cell.fill = row_fill

        ws.cell(row=r_idx, column=15).alignment = WRAP_ALIGN

    _auto_col_width(ws)
    ws.column_dimensions[get_column_letter(15)].width = 50
    ws.row_dimensions[1].height = 20

    wb.save(output_path)
    print(f'  已輸出: {output_path}（{len(duplicate_list)} 筆重複記錄）')


# ============================================================
# v2 — 4. 人工判讀紀錄範本
# ============================================================
REVIEW_HEADERS = [
    '來源CSV', '原始列號', '原始識別碼', '管線識別碼', '點號',
    'X', 'Y', 'Z', '測量日期',
    '判讀狀態', '問題類型', '判讀說明', '建議回覆施工廠商內容',
    '確認人員', '確認日期',
]
# 可填寫欄位從第 10 欄（J）起
REVIEW_EDITABLE_START_COL = 10

REVIEW_STATUS_LIST  = '正常,待確認,需廠商補測,排除,採用'
PROBLEM_TYPE_LIST   = '點位偏移,點位缺漏,點位重複,座標異常,日期不一致,點序錯誤,其他'


def _add_dropdown(ws, col_letter, start_row, end_row, formula, prompt_title='', prompt=''):
    """在指定欄位範圍加入下拉選單驗證。"""
    kwargs = dict(
        type='list',
        formula1=f'"{formula}"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle='無效選項',
        error='請從下拉選單中選擇，或清空此欄。',
    )
    if prompt_title:
        kwargs['showInputMessage'] = True
        kwargs['promptTitle'] = prompt_title
        kwargs['prompt'] = prompt
    dv = DataValidation(**kwargs)
    dv.add(f'{col_letter}{start_row}:{col_letter}{end_row}')
    ws.add_data_validation(dv)


def export_review_template(df, anomaly_set, output_path):
    """
    產生人工判讀紀錄範本。
    anomaly_set: set of (source_file, original_row_int) 有異常的列
    異常列預填「待確認」，其餘留白，供人工逐點填寫。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '人工判讀紀錄'
    ws.freeze_panes = 'A2'

    # ── 標題列 ──────────────────────────────
    for col_idx, h in enumerate(REVIEW_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL_GREEN
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # ── 資料列 ──────────────────────────────
    # 依管線識別碼 + 點號整數排序
    df_sorted = df.copy()
    df_sorted['_sort_pid'] = df_sorted['pipeline_id'].fillna('')
    df_sorted['_sort_pno'] = df_sorted['point_no_int'].apply(
        lambda v: int(v) if (v is not None and not (isinstance(v, float) and math.isnan(v))) else 99999
    )
    df_sorted = df_sorted.sort_values(['_sort_pid', '_sort_pno'])

    for r_idx, (_, row) in enumerate(df_sorted.iterrows(), start=2):
        src = row.get('_source_file', '')
        orig_row = row.get('_original_row')
        is_anomaly = (src, orig_row) in anomaly_set if orig_row else False

        values = [
            src,
            orig_row,
            row.get('raw_id'),
            row.get('pipeline_id'),
            row.get('point_no'),
            row.get('raw_x'),
            row.get('raw_y'),
            row.get('raw_z'),
            row.get('raw_date'),
            '待確認' if is_anomaly else '',   # 判讀狀態（預填）
            '',  # 問題類型
            '',  # 判讀說明
            '',  # 建議回覆施工廠商內容
            '',  # 確認人員
            '',  # 確認日期
        ]

        for c_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=_fmt(val))
            cell.border = THIN_BORDER

            is_editable = (c_idx >= REVIEW_EDITABLE_START_COL)
            if is_editable:
                # 可填寫欄位：淡藍底色
                cell.font = BODY_FONT
                cell.fill = EDITABLE_COL_FILL
                cell.alignment = LEFT_ALIGN
            elif is_anomaly:
                # 有異常的資料列：淡黃底色
                cell.font = BODY_FONT
                cell.fill = ANOMALY_ROW_FILL
                cell.alignment = LEFT_ALIGN
            else:
                cell.font = BODY_FONT
                cell.alignment = LEFT_ALIGN

    last_data_row = len(df_sorted) + 1
    max_dv_row = max(last_data_row, 1001)  # 下拉選單至少覆蓋到第 1001 列

    # ── 下拉選單 ────────────────────────────
    _add_dropdown(ws, 'J', 2, max_dv_row, REVIEW_STATUS_LIST,
                  prompt_title='判讀狀態', prompt='請選擇：正常 / 待確認 / 需廠商補測 / 排除 / 採用')
    _add_dropdown(ws, 'K', 2, max_dv_row, PROBLEM_TYPE_LIST,
                  prompt_title='問題類型', prompt='請選擇問題類型（無問題可留空）')

    # ── 欄寬 ────────────────────────────────
    _auto_col_width(ws, min_w=8, max_w=35)
    ws.column_dimensions['M'].width = 40  # 建議回覆施工廠商內容
    ws.column_dimensions['L'].width = 30  # 判讀說明
    ws.row_dimensions[1].height = 22

    # ── 說明備註（首列上方插入一列說明）───────
    ws.insert_rows(1)
    note_cell = ws.cell(row=1, column=1,
                        value='【人工判讀紀錄範本】淡藍色欄位請人工填寫。淡黃色列為程式偵測到的異常點，已預填「待確認」。本範本不會修改原始 CSV。')
    note_cell.font = Font(name='微軟正黑體', size=10, bold=True, color='1A5632')
    note_cell.alignment = LEFT_ALIGN
    ws.merge_cells(f'A1:{get_column_letter(len(REVIEW_HEADERS))}1')
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = 'A3'  # 凍結說明列 + 標題列

    wb.save(output_path)
    print(f'  已輸出: {output_path}（{len(df_sorted)} 筆）')


# ============================================================
# v2 — 5. 修正紀錄範本
# ============================================================
CORRECTION_HEADERS = [
    '動作類型',
    '原始識別碼', '管線識別碼', '點號',
    '原X', '原Y', '原Z',
    '修正後X', '修正後Y', '修正後Z',
    '修正原因', '是否納入後續GML',
    '確認人員', '確認日期',
]

ACTION_TYPE_LIST = '新增,刪除,採用,排除,修正'
YES_NO_LIST      = '是,否'

BLANK_ROWS = 300   # 空白列數（供人工填寫）


def export_correction_template(output_path):
    """產生修正紀錄範本（空白，供人工逐筆填寫修正動作）。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '修正紀錄'
    ws.freeze_panes = 'A3'

    # ── 說明列（第 1 列）────────────────────
    note_cell = ws.cell(row=1, column=1,
                        value='【修正紀錄範本】每筆填寫一個修正動作。本範本不會修改原始 CSV，僅供人工記錄確認後的修正決策。')
    note_cell.font = Font(name='微軟正黑體', size=10, bold=True, color='4A235A')
    note_cell.alignment = LEFT_ALIGN
    ws.merge_cells(f'A1:{get_column_letter(len(CORRECTION_HEADERS))}1')
    ws.row_dimensions[1].height = 18

    # ── 標題列（第 2 列）────────────────────
    for col_idx, h in enumerate(CORRECTION_HEADERS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL_PURPLE
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[2].height = 22

    # ── 空白資料列（第 3 列起）──────────────
    last_dv_row = BLANK_ROWS + 2
    for r_idx in range(3, last_dv_row + 1):
        for c_idx in range(1, len(CORRECTION_HEADERS) + 1):
            cell = ws.cell(row=r_idx, column=c_idx, value='')
            cell.border = THIN_BORDER
            cell.font = BODY_FONT
            cell.fill = EDITABLE_COL_FILL
            cell.alignment = LEFT_ALIGN

    # ── 下拉選單 ────────────────────────────
    _add_dropdown(ws, 'A', 3, last_dv_row, ACTION_TYPE_LIST,
                  prompt_title='動作類型', prompt='新增 / 刪除 / 採用 / 排除 / 修正')
    _add_dropdown(ws, 'L', 3, last_dv_row, YES_NO_LIST,
                  prompt_title='是否納入後續GML', prompt='是 / 否')

    # ── 欄寬 ────────────────────────────────
    col_widths = {
        'A': 12,   # 動作類型
        'B': 28,   # 原始識別碼
        'C': 28,   # 管線識別碼
        'D': 8,    # 點號
        'E': 14,   # 原X
        'F': 14,   # 原Y
        'G': 12,   # 原Z
        'H': 14,   # 修正後X
        'I': 14,   # 修正後Y
        'J': 12,   # 修正後Z
        'K': 35,   # 修正原因
        'L': 18,   # 是否納入後續GML
        'M': 14,   # 確認人員
        'N': 14,   # 確認日期
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(output_path)
    print(f'  已輸出: {output_path}（{BLANK_ROWS} 列空白，可直接填寫）')


# ============================================================
# 統一入口
# ============================================================
def export_all(df, anomaly_list, duplicate_list, output_dir):
    os.makedirs(output_dir, exist_ok=True)
    export_parsed_points(df, os.path.join(output_dir, '解析後點位表.xlsx'))
    export_anomalies(anomaly_list,   os.path.join(output_dir, '異常清單.xlsx'))
    export_duplicates(duplicate_list, os.path.join(output_dir, '重複量測比對表.xlsx'))

    # v2 新增
    anomaly_set = set()
    for item in anomaly_list:
        src = item.get('來源CSV')
        row_no = item.get('原始列號')
        if src and row_no and str(row_no) != '（整份檔案）':
            try:
                anomaly_set.add((src, int(row_no)))
            except (ValueError, TypeError):
                pass
    export_review_template(df, anomaly_set, os.path.join(output_dir, '人工判讀紀錄範本.xlsx'))
    export_correction_template(os.path.join(output_dir, '修正紀錄範本.xlsx'))


# ============================================================
# Streamlit 輔助：BytesIO 輸出函式
# ============================================================

def _wb_to_bytes(wb):
    """將 openpyxl Workbook 儲存到 BytesIO 並回傳。"""
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _export_to_bytes(export_func, *args):
    """
    呼叫任意 export_XXX(... , output_path) 函式，
    輸出至暫存檔後讀回 BytesIO，適用於 Streamlit 下載按鈕。
    export_func 的 output_path 必須是最後一個位置引數。
    """
    fd, tmppath = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)
    try:
        export_func(*args, tmppath)
        with open(tmppath, 'rb') as f:
            buf = _io.BytesIO(f.read())
        buf.seek(0)
        return buf
    finally:
        try:
            os.unlink(tmppath)
        except Exception:
            pass


def export_review_df_bytes(review_df):
    """
    將 Streamlit data_editor 回傳的人工判讀 DataFrame
    格式化輸出為 Excel BytesIO（供下載按鈕使用）。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '人工判讀紀錄'
    ws.freeze_panes = 'A3'

    # 說明列
    note = ws.cell(row=1, column=1,
                   value='【人工判讀紀錄】由 Streamlit 網頁版匯出。淡藍色為可填寫欄位。')
    note.font = Font(name='微軟正黑體', size=10, bold=True, color='1A5632')
    note.alignment = LEFT_ALIGN
    ws.merge_cells(f'A1:{get_column_letter(len(REVIEW_HEADERS))}1')
    ws.row_dimensions[1].height = 18

    # 標題列
    for col_idx, h in enumerate(REVIEW_HEADERS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL_GREEN
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[2].height = 22

    # 資料列
    for r_idx, (_, row) in enumerate(review_df.iterrows(), start=3):
        for c_idx, col_name in enumerate(REVIEW_HEADERS, start=1):
            val = row.get(col_name) if col_name in review_df.columns else ''
            cell = ws.cell(row=r_idx, column=c_idx, value=_fmt(val))
            cell.border = THIN_BORDER
            cell.font = BODY_FONT
            if c_idx >= REVIEW_EDITABLE_START_COL:
                cell.fill = EDITABLE_COL_FILL
            cell.alignment = LEFT_ALIGN

    last_row = max(len(review_df) + 2, 1001)
    _add_dropdown(ws, 'J', 3, last_row, REVIEW_STATUS_LIST)
    _add_dropdown(ws, 'K', 3, last_row, PROBLEM_TYPE_LIST)

    _auto_col_width(ws, min_w=8, max_w=35)
    ws.column_dimensions['M'].width = 40
    ws.column_dimensions['L'].width = 28

    return _wb_to_bytes(wb)


def export_correction_df_bytes(correction_df):
    """
    將 Streamlit data_editor 回傳的修正紀錄 DataFrame
    格式化輸出為 Excel BytesIO（供下載按鈕使用）。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '修正紀錄'
    ws.freeze_panes = 'A3'

    note = ws.cell(row=1, column=1,
                   value='【修正紀錄】由 Streamlit 網頁版匯出。每列代表一個修正動作。')
    note.font = Font(name='微軟正黑體', size=10, bold=True, color='4A235A')
    note.alignment = LEFT_ALIGN
    ws.merge_cells(f'A1:{get_column_letter(len(CORRECTION_HEADERS))}1')
    ws.row_dimensions[1].height = 18

    for col_idx, h in enumerate(CORRECTION_HEADERS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL_PURPLE
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[2].height = 22

    for r_idx, (_, row) in enumerate(correction_df.iterrows(), start=3):
        for c_idx, col_name in enumerate(CORRECTION_HEADERS, start=1):
            val = row.get(col_name) if col_name in correction_df.columns else ''
            cell = ws.cell(row=r_idx, column=c_idx, value=_fmt(val))
            cell.border = THIN_BORDER
            cell.font = BODY_FONT
            cell.fill = EDITABLE_COL_FILL
            cell.alignment = LEFT_ALIGN

    last_row = max(len(correction_df) + 2, 302)
    _add_dropdown(ws, 'A', 3, last_row, ACTION_TYPE_LIST)
    _add_dropdown(ws, 'L', 3, last_row, YES_NO_LIST)

    col_widths = {
        'A': 12, 'B': 28, 'C': 28, 'D': 8,
        'E': 14, 'F': 14, 'G': 12,
        'H': 14, 'I': 14, 'J': 12,
        'K': 35, 'L': 18, 'M': 14, 'N': 14,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    return _wb_to_bytes(wb)
