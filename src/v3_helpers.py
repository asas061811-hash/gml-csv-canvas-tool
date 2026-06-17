"""
v3_helpers.py - 第三版輔助模組

提供：
  - 可編輯點位表 (edited_points_df) 初始化
  - v3 版本 Plotly 畫布建構（支援點選、per-point 樣式）
  - 修正紀錄表 (edit_log_df) 管理
  - 多工作表 Excel 匯出
"""

import datetime
import io
import json
import math

import openpyxl
import pandas as pd
import plotly.graph_objects as go
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── 顏色色票 ─────────────────────────────────────────
COLOR_PALETTE = [
    '#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd',
    '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22', '#17becf',
    '#aec7e8', '#ffbb78', '#98df8a', '#ff9896', '#c5b0d5',
    '#c49c94', '#f7b6d2', '#c7c7c7', '#dbdb8d', '#9edae5',
]

# ── 選項清單（供 app.py import）────────────────────────
REVIEW_STATUS_OPTIONS  = ['未判讀', '正常', '待確認', '需廠商補測', '排除', '採用', '已修正']
PROBLEM_TYPE_OPTIONS   = ['無', '點位偏移', '點位缺漏', '點位重複', '座標異常', '日期不一致', '點序錯誤', '手動排除', '其他']
INCLUDE_RESULT_OPTIONS = ['是', '否']

# ── 管線分類對照表（公共設施管線資料庫）──────────────
_MAJOR_NAMES = {
    '8': '公共設施管線資料庫',
}

_MID_NAMES = {
    '01': '電信',
    '02': '電力',
    '03': '自來水',
    '04': '下水道',
    '05': '瓦斯',
    '06': '水利',
    '07': '輸油',
    '08': '綜合',
}

_MINOR_NAMES = {
    '01': {'01': '一般電信系統', '02': '軍訊系統', '03': '警訊系統',
           '04': '有線電視系統', '05': '交通號誌系統'},
    '02': {'01': '配電系統', '02': '路燈電力系統', '03': '交通號誌電力系統', '04': '輸電系統'},
    '03': {'01': '自來水系統'},
    '04': {'01': '污水系統', '02': '雨水系統', '03': '合流系統'},
    '05': {'01': '供氣系統'},
    '06': {'01': '灌排系統'},
    '07': {'01': '輸油系統'},
    '08': {'01': '共同管道', '02': '寬頻管道'},
}

_FINE_NAMES_DEFAULT = {
    '01': '管線', '02': '人手孔', '03': '電桿', '04': '開關',
    '96': '其他設施', '97': '場站',
}

_FINE_NAMES = {
    '06': {'01': '管線', '02': '閘門', '96': '其他設施', '97': '場站'},
    '08': {'01': '管線', '02': '人手孔', '03': '維護口', '96': '其他設施', '97': '場站'},
}

# ── 修正紀錄表欄位 ────────────────────────────────────
EDIT_LOG_COLS = [
    '修正時間', 'point_key', '動作類型',
    '原始_來源CSV', '原始_列號', '原始_識別碼',
    '修改前_資料', '修改後_資料',
    '判讀狀態', '問題類型', '人工備註', '是否納入成果',
]

# ── 可追溯的原始欄位（使用者可修改顯示值，但這些欄位永不改變）──
_IMMUTABLE_COLS = [
    '_source_file', '_original_row', 'raw_id',
    'pipeline_id', 'point_no', 'point_no_int',
    'raw_category', 'raw_x', 'raw_y', 'raw_z', 'raw_date',
    'x_val', 'y_val', 'z_val',
]


# ============================================================
# 工具函式
# ============================================================

def make_point_key(source_file, original_row, raw_id):
    """產生點位唯一識別碼（不受使用者編輯影響）。"""
    return f"{source_file or ''}||{original_row or ''}||{raw_id or ''}"


def safe_float(val):
    if val is None or str(val).strip() == '':
        return None
    try:
        v = float(str(val).strip())
        return None if math.isnan(v) else v
    except (ValueError, TypeError):
        return None


def _now_str():
    return datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')


def _try_int_pno(pno_str):
    try:
        return int(str(pno_str).strip()) if pno_str else None
    except (ValueError, TypeError):
        return None


def decode_category(code):
    """
    解析 7 碼類別碼（格式 XYYZZWW）。
    回傳 dict：major/mid/minor/fine 的 code 與 name，以及 parse_ok 旗標。
    無法解析時各 name 填 '未知'，parse_ok=False。
    """
    empty = {
        'major_code': '', 'major_name': '未知',
        'mid_code':   '', 'mid_name':   '未知',
        'minor_code': '', 'minor_name': '未知',
        'fine_code':  '', 'fine_name':  '未知',
        'parse_ok': False,
    }
    if not code or str(code).strip() == '':
        return empty
    s = str(code).strip()
    if len(s) != 7 or not s.isdigit():
        return empty

    major_c = s[0]
    mid_c   = s[1:3]
    minor_c = s[3:5]
    fine_c  = s[5:7]

    fine_map = _FINE_NAMES.get(mid_c, _FINE_NAMES_DEFAULT)

    return {
        'major_code': major_c,
        'major_name': _MAJOR_NAMES.get(major_c, '未知'),
        'mid_code':   mid_c,
        'mid_name':   _MID_NAMES.get(mid_c, '未知'),
        'minor_code': minor_c,
        'minor_name': _MINOR_NAMES.get(mid_c, {}).get(minor_c, '未知'),
        'fine_code':  fine_c,
        'fine_name':  fine_map.get(fine_c, '未知'),
        'parse_ok': True,
    }


def count_pipelines(df, pid_col='edited_pipeline_id'):
    """
    統計唯一管線數。
    - 排除空字串、None、NaN；同一管線識別碼有多個點位仍只算 1 條。
    - 若欄位不存在則嘗試 'pipeline_id'（原始欄）。
    - 此函式為全系統統一管線計數入口，避免各處計算不一致。
    """
    if df is None or df.empty:
        return 0
    col = pid_col if pid_col in df.columns else (
        'pipeline_id' if 'pipeline_id' in df.columns else None
    )
    if col is None:
        return 0
    return int(df[col].replace('', pd.NA).dropna().nunique())


# ============================================================
# 設施分類表（表 6-9 ～ 表 6-17）
# ============================================================

def _vstr(val):
    """None / NaN → ''，其他轉 str（v3_helpers 內部用）。"""
    if val is None:
        return ''
    try:
        if math.isnan(float(val)):
            return ''
    except (TypeError, ValueError):
        pass
    return str(val)


FACILITY_TYPES = ['管線', '人手孔', '開關閥', '消防栓', '電桿', '號誌', '其他設施', '維護口', '場站']


def get_facility_type(category_code):
    """
    依類別碼最後兩碼（細類碼）判斷設施類型（表 6-9 ～ 表 6-17）。
    特例：水利(mid=06) fine=02 → 開關閥；綜合(mid=08) fine=03 → 維護口。
    """
    info = decode_category(category_code)
    if not info['parse_ok']:
        return '其他設施'
    mid  = info['mid_code']
    fine = info['fine_code']
    if fine == '01':
        return '管線'
    if fine == '02':
        return '開關閥' if mid == '06' else '人手孔'
    if fine == '03':
        return '維護口' if mid == '08' else '電桿'
    if fine == '04':
        return '開關閥'
    if fine == '05':
        return '消防栓'
    if fine == '06':
        return '號誌'
    if fine == '96':
        return '其他設施'
    if fine == '97':
        return '場站'
    return '其他設施'


_PIPELINE_BASE_COLS = [
    '類別碼', '識別碼', '起點編號', '終點編號', '管理單位', '作業區分', '設置日期',
    '管線編號', '尺寸單位', '管徑寬度', '管徑高度', '涵管條數', '管線材料',
    '起點埋設深度', '終點埋設深度', '管線長度', '管線型態', '使用狀態', '資料狀態',
    '備註', '壓力區分', '輸送物質', '管線識別碼', '點位數量',
]
_PIPELINE_TAIL_COLS = ['來源CSV', '判讀狀態', '問題類型', '人工備註']

FACILITY_SCHEMAS = {
    '管線': None,  # 動態欄位（橫向展開點位座標），由 build_classification_df 處理
    '人手孔': [
        '類別碼', '識別碼', '管理單位', '作業區分', '設置日期', '人手孔編號',
        '孔蓋種類', '尺寸單位', '蓋部寬度', '蓋部長度', '閘門名稱', '地盤高',
        '孔深', '孔蓋型態', '使用狀態', '資料狀態', '內容物', '備註',
        'X', 'Y', 'Z', 'd', '來源CSV', '原始列號', '判讀狀態', '問題類型', '人工備註',
    ],
    '開關閥': [
        '類別碼', '識別碼', '管理單位', '作業區分', '設置日期', '開關閥編號',
        '閥類編號', '口徑', '名稱', '地盤高', '埋設深度', '開關閥型態',
        '使用狀態', '資料狀態', '備註',
        'X', 'Y', 'Z', 'd', '來源CSV', '原始列號', '判讀狀態', '問題類型', '人工備註',
    ],
    '消防栓': [
        '類別碼', '識別碼', '管理單位', '作業區分', '設置日期', '消防栓編號',
        '管身口徑', '出水口口徑', '埋設深度', '消防栓型態', '使用狀態', '資料狀態', '備註',
        'X', 'Y', 'Z', 'd', '來源CSV', '原始列號', '判讀狀態', '問題類型', '人工備註',
    ],
    '電桿': [
        '類別碼', '識別碼', '管理單位', '作業區分', '設置日期', '電桿編號',
        '長度', '材質', '使用狀態', '資料狀態', '備註',
        'X', 'Y', 'Z', 'd', '來源CSV', '原始列號', '判讀狀態', '問題類型', '人工備註',
    ],
    '號誌': [
        '類別碼', '識別碼', '管理單位', '作業區分', '設置日期', '號誌編號',
        '號誌種類', '號誌架設方式', '長度', '使用狀態', '資料狀態', '備註',
        'X', 'Y', 'Z', 'd', '來源CSV', '原始列號', '判讀狀態', '問題類型', '人工備註',
    ],
    '其他設施': [
        '類別碼', '識別碼', '管理單位', '作業區分', '設置日期', '設施編號',
        '設施名稱', '設施長度', '設施寬度', '設施高度', '設施型態',
        '使用狀態', '資料狀態', '備註',
        'X', 'Y', 'Z', 'd', '來源CSV', '原始列號', '判讀狀態', '問題類型', '人工備註',
    ],
    '維護口': [
        '類別碼', '識別碼', '管理單位', '作業區分', '設置日期', '維護口編號',
        '名稱', '使用狀態', '資料狀態', '備註',
        'X', 'Y', 'Z', 'd', '來源CSV', '原始列號', '判讀狀態', '問題類型', '人工備註',
    ],
    '場站': [
        '類別碼', '識別碼', '管理單位', '作業區分', '設置日期', '場站名稱',
        '使用狀態', '資料狀態', '備註',
        'X', 'Y', 'Z', 'd', '來源CSV', '原始列號', '判讀狀態', '問題類型', '人工備註',
    ],
}

_CLS_FIELD_MAP = {
    '類別碼':     'edited_category',
    '識別碼':     'edited_raw_id',
    '管線識別碼':  'edited_pipeline_id',
    'X':          'edited_x',
    'Y':          'edited_y',
    'Z':          'edited_z',
    '來源CSV':    '_source_file',
    '原始列號':   '_original_row',
    '判讀狀態':   'review_status',
    '問題類型':   'problem_type',
    '人工備註':   'manual_note',
}


def _facility_filter(edited_df, facility_type):
    """篩選出指定設施類型的列。"""
    def _ftype(c):
        s = str(c).strip() if (c is not None and not (isinstance(c, float) and math.isnan(c))) else ''
        return get_facility_type(s)
    return edited_df[edited_df['edited_category'].apply(_ftype) == facility_type].copy()


def _row_to_cls_dict(row, schema_cols):
    """將 edited_df 的一列依 schema_cols 轉換為 dict，缺少來源欄位留空字串。"""
    d = {col: '' for col in schema_cols}
    for col in schema_cols:
        src = _CLS_FIELD_MAP.get(col)
        if src:
            d[col] = _vstr(row.get(src))
    return d


def _build_pipeline_classification(fdf):
    """管線：依 edited_pipeline_id 群組，每條管線一列，點位座標橫向展開。"""
    if fdf.empty:
        return pd.DataFrame(columns=_PIPELINE_BASE_COLS + _PIPELINE_TAIL_COLS)

    def _pno_key(r):
        v = r.get('edited_point_no_int')
        try:
            if v is None or pd.isna(v):
                return (1, 0)
            return (0, int(v))
        except Exception:
            return (1, 0)

    grouped = {}
    for _, row in fdf.iterrows():
        pid = _vstr(row.get('edited_pipeline_id'))
        key = pid if pid else f'__nopid__{row["point_key"]}'
        grouped.setdefault(key, []).append(row)

    max_pts = 0
    pipeline_rows = []
    for key, pts in grouped.items():
        pts_sorted = sorted(pts, key=_pno_key)
        max_pts    = max(max_pts, len(pts_sorted))
        first      = pts_sorted[0]

        base = {col: '' for col in _PIPELINE_BASE_COLS}
        base['類別碼']     = _vstr(first.get('edited_category'))
        base['識別碼']     = _vstr(first.get('edited_raw_id'))
        base['管線識別碼'] = _vstr(first.get('edited_pipeline_id'))
        base['點位數量']   = str(len(pts_sorted))

        tail = {col: '' for col in _PIPELINE_TAIL_COLS}
        tail['來源CSV']  = _vstr(first.get('_source_file'))
        tail['判讀狀態'] = _vstr(first.get('review_status'))
        tail['問題類型'] = _vstr(first.get('problem_type'))
        tail['人工備註'] = _vstr(first.get('manual_note'))

        pt_data = {}
        for n, pt in enumerate(pts_sorted, 1):
            pt_data[f'第{n}點X'] = _vstr(pt.get('edited_x'))
            pt_data[f'第{n}點Y'] = _vstr(pt.get('edited_y'))
            pt_data[f'第{n}點Z'] = _vstr(pt.get('edited_z'))
            pt_data[f'第{n}點d'] = ''

        pipeline_rows.append({**base, **pt_data, **tail})

    all_cols = list(_PIPELINE_BASE_COLS)
    for n in range(1, max_pts + 1):
        all_cols += [f'第{n}點X', f'第{n}點Y', f'第{n}點Z', f'第{n}點d']
    all_cols += list(_PIPELINE_TAIL_COLS)

    df = pd.DataFrame(pipeline_rows)
    for col in all_cols:
        if col not in df.columns:
            df[col] = ''
    return df[all_cols].fillna('')


def build_classification_df(edited_df, facility_type, include_excluded=False):
    """
    依設施類型（表 6-9 ～ 表 6-17）建立分類表 DataFrame。
    - 管線：一條管線一列，點位座標橫向展開（第N點X/Y/Z/d）。
    - 其他：一個點位一列，固定欄位，缺少的來源欄位留空。
    - include_excluded=False（預設）：已排除點位不出現在分類表。
    """
    if facility_type not in FACILITY_TYPES:
        return pd.DataFrame()
    if 'edited_category' not in edited_df.columns:
        schema = FACILITY_SCHEMAS.get(facility_type)
        return pd.DataFrame(columns=schema or _PIPELINE_BASE_COLS + _PIPELINE_TAIL_COLS)

    # 過濾已排除點位（預設行為：排除不納入成果的點位）
    if not include_excluded:
        edited_df = edited_df[~edited_df.apply(_is_excluded, axis=1)].copy()

    fdf = _facility_filter(edited_df, facility_type)

    if facility_type == '管線':
        return _build_pipeline_classification(fdf)

    schema = FACILITY_SCHEMAS[facility_type]
    if fdf.empty:
        return pd.DataFrame(columns=schema)
    rows = [_row_to_cls_dict(r, schema) for _, r in fdf.iterrows()]
    return pd.DataFrame(rows, columns=schema)


def build_cls_pkey_list(edited_df, facility_type, include_excluded=False):
    """
    回傳與 build_classification_df 同順序的 point_key 清單。
    - 點位型：每個元素為一個 point_key 字串。
    - 管線型：每個元素為該管線包含的 point_key 字串 list。
    用於分類表勾選排除時，能正確對應到 edited_points_df 的點位。
    """
    if 'edited_category' not in edited_df.columns or 'point_key' not in edited_df.columns:
        return []
    if not include_excluded:
        edited_df = edited_df[~edited_df.apply(_is_excluded, axis=1)].copy()
    fdf = _facility_filter(edited_df, facility_type)
    if fdf.empty:
        return []
    if facility_type != '管線':
        return [_vstr(r.get('point_key')) for _, r in fdf.iterrows()]
    # 管線：依相同群組邏輯
    grouped = {}
    for _, row in fdf.iterrows():
        pid = _vstr(row.get('edited_pipeline_id'))
        key = pid if pid else f'__nopid__{row["point_key"]}'
        grouped.setdefault(key, []).append(row)
    result = []
    for pts in grouped.values():
        result.append([_vstr(pt.get('point_key')) for pt in pts])
    return result


def _row_snapshot(row_dict, fields):
    """從 row dict 取出指定欄位，產生 JSON 字串快照。"""
    snap = {}
    for f in fields:
        v = row_dict.get(f)
        if isinstance(v, float) and math.isnan(v):
            v = None
        snap[f] = v
    return json.dumps(snap, ensure_ascii=False, default=str)


_SNAP_FIELDS = [
    'edited_category', 'edited_raw_id',
    'edited_source_file', 'edited_original_row',
    'edited_pipeline_id', 'edited_point_no',
    'edited_x', 'edited_y', 'edited_z', 'edited_date',
    'review_status', 'problem_type', 'manual_note', 'include_in_result',
]


# ============================================================
# 初始化資料表
# ============================================================

def init_edited_points_df(df):
    """
    從原始解析 DataFrame 建立可編輯點位表。
    初始值等於原始值；所有 edited_* 欄位供人工修正。
    原始欄位（raw_*、_source_file 等）永遠保留，不受編輯影響。
    """
    edf = df.copy()

    # 唯一 key（永不改變）
    edf['point_key'] = edf.apply(
        lambda r: make_point_key(r['_source_file'], r['_original_row'], r.get('raw_id')),
        axis=1,
    )

    # 可編輯顯示欄位（初始 = 原始解析值）
    # 全部轉成 object dtype，避免 int64/float64 欄位在 .loc[] 設字串值時報 TypeError
    edf['edited_category']      = edf['raw_category'].astype(object)
    edf['edited_raw_id']        = edf['raw_id'].astype(object)
    edf['edited_source_file']   = edf['_source_file'].astype(object)
    edf['edited_original_row']  = edf['_original_row'].astype(object)
    edf['edited_pipeline_id']   = edf['pipeline_id'].astype(object)
    edf['edited_point_no']      = edf['point_no'].astype(object)
    edf['edited_point_no_int']  = edf['point_no_int']            # 保留 Int64 供排序
    edf['edited_x']             = edf['raw_x'].astype(object)
    edf['edited_y']             = edf['raw_y'].astype(object)
    edf['edited_z']             = edf['raw_z'].astype(object)
    edf['edited_date']          = edf['raw_date'].astype(object)
    edf['edited_x_val']         = edf['x_val']
    edf['edited_y_val']         = edf['y_val']
    edf['edited_z_val']         = edf['z_val']

    # 判讀欄位
    edf['review_status']     = '未判讀'
    edf['problem_type']      = '無'
    edf['manual_note']       = ''
    edf['include_in_result'] = '是'
    edf['is_modified']       = False

    return edf


def init_edit_log():
    return pd.DataFrame(columns=EDIT_LOG_COLS)


# ============================================================
# 修正 / 還原 / 快速標記操作
# ============================================================

def _save_log(edit_log_df, point_key, action, orig_info, before_snap, after_snap,
              review_status, problem_type, manual_note, include_in_result):
    """建立一筆修正紀錄並附加到 edit_log_df。"""
    row = {
        '修正時間':    _now_str(),
        'point_key':  point_key,
        '動作類型':   action,
        '原始_來源CSV':  orig_info['_source_file'],
        '原始_列號':     orig_info['_original_row'],
        '原始_識別碼':   orig_info.get('raw_id'),
        '修改前_資料':   before_snap,
        '修改後_資料':   after_snap,
        '判讀狀態':   review_status,
        '問題類型':   problem_type,
        '人工備註':   manual_note,
        '是否納入成果': include_in_result,
    }
    return pd.concat([edit_log_df, pd.DataFrame([row])], ignore_index=True)


def save_point_edit(edited_df, edit_log_df, point_key, form_data):
    """
    以 form_data 更新 edited_df 中指定點位，並寫入修正紀錄。
    form_data 鍵：category, raw_id, source_file, original_row,
                  pipeline_id, point_no, x, y, z, date,
                  review_status, problem_type, manual_note, include_in_result,
                  action_type
    回傳 (updated_edited_df, updated_edit_log_df)
    """
    edited_df = edited_df.copy()
    idxs = edited_df.index[edited_df['point_key'] == point_key].tolist()
    if not idxs:
        return edited_df, edit_log_df

    i = idxs[0]
    prev = edited_df.loc[i].to_dict()
    before_snap = _row_snapshot(prev, _SNAP_FIELDS)

    def _s(key):
        v = form_data.get(key)
        return str(v).strip() if v is not None else None

    new_cat    = _s('category') or None
    new_rid    = _s('raw_id') or None
    new_src    = _s('source_file') or None
    new_rrow   = _s('original_row') or None
    new_pid    = _s('pipeline_id') or None
    new_pno    = _s('point_no') or None
    new_x      = _s('x') or None
    new_y      = _s('y') or None
    new_z      = _s('z') or None
    new_date   = _s('date') or None
    new_status = form_data.get('review_status', '未判讀')
    new_prob   = form_data.get('problem_type', '無')
    new_note   = str(form_data.get('manual_note', '') or '')
    new_incl   = form_data.get('include_in_result', '是')
    action     = form_data.get('action_type', '修正')

    edited_df.loc[i, 'edited_category']     = new_cat
    edited_df.loc[i, 'edited_raw_id']       = new_rid
    edited_df.loc[i, 'edited_source_file']  = new_src
    edited_df.loc[i, 'edited_original_row'] = new_rrow
    edited_df.loc[i, 'edited_pipeline_id']  = new_pid
    edited_df.loc[i, 'edited_point_no']     = new_pno
    edited_df.loc[i, 'edited_point_no_int'] = pd.array(
        [_try_int_pno(new_pno)], dtype=pd.Int64Dtype())[0]
    edited_df.loc[i, 'edited_x']            = new_x
    edited_df.loc[i, 'edited_y']            = new_y
    edited_df.loc[i, 'edited_z']            = new_z
    edited_df.loc[i, 'edited_date']         = new_date
    edited_df.loc[i, 'edited_x_val']        = safe_float(new_x)
    edited_df.loc[i, 'edited_y_val']        = safe_float(new_y)
    edited_df.loc[i, 'edited_z_val']        = safe_float(new_z)
    edited_df.loc[i, 'review_status']       = new_status
    edited_df.loc[i, 'problem_type']        = new_prob
    edited_df.loc[i, 'manual_note']         = new_note
    edited_df.loc[i, 'include_in_result']   = new_incl
    edited_df.loc[i, 'is_modified']         = True

    after_snap = _row_snapshot(edited_df.loc[i].to_dict(), _SNAP_FIELDS)

    # 原始追溯資訊（不受編輯影響）
    orig_info = {
        '_source_file': prev.get('_source_file'),
        '_original_row': prev.get('_original_row'),
        'raw_id': prev.get('raw_id'),
    }
    edit_log_df = _save_log(
        edit_log_df, point_key, action, orig_info,
        before_snap, after_snap,
        new_status, new_prob, new_note, new_incl,
    )
    return edited_df, edit_log_df


def quick_mark(edited_df, edit_log_df, point_key, action):
    """
    快速標記：action 為 '標記正常' 或 '標記排除'。
    不修改座標，只更新判讀狀態與是否納入成果。
    """
    edited_df = edited_df.copy()
    idxs = edited_df.index[edited_df['point_key'] == point_key].tolist()
    if not idxs:
        return edited_df, edit_log_df

    i = idxs[0]
    prev = edited_df.loc[i].to_dict()
    before_snap = _row_snapshot(prev, _SNAP_FIELDS)

    if action == '標記正常':
        new_status = '正常'
        new_incl   = '是'
    else:  # 標記排除
        new_status = '排除'
        new_incl   = '否'

    edited_df.loc[i, 'review_status']     = new_status
    edited_df.loc[i, 'include_in_result'] = new_incl
    edited_df.loc[i, 'is_modified']       = True

    after_snap = _row_snapshot(edited_df.loc[i].to_dict(), _SNAP_FIELDS)
    orig_info  = {
        '_source_file':  prev.get('_source_file'),
        '_original_row': prev.get('_original_row'),
        'raw_id':        prev.get('raw_id'),
    }
    edit_log_df = _save_log(
        edit_log_df, point_key, action, orig_info,
        before_snap, after_snap,
        new_status, prev.get('problem_type', '無'),
        prev.get('manual_note', ''), new_incl,
    )
    return edited_df, edit_log_df


def restore_point(edited_df, edit_log_df, point_key, original_df):
    """將指定點位的所有 edited_* 欄位還原為原始解析值，並寫入還原紀錄。"""
    edited_df = edited_df.copy()
    idxs = edited_df.index[edited_df['point_key'] == point_key].tolist()
    if not idxs:
        return edited_df, edit_log_df

    i = idxs[0]
    curr = edited_df.loc[i].to_dict()
    before_snap = _row_snapshot(curr, _SNAP_FIELDS)

    mask = (
        (original_df['_source_file'] == curr['_source_file']) &
        (original_df['_original_row'] == curr['_original_row'])
    )
    orig_rows = original_df[mask]
    if orig_rows.empty:
        return edited_df, edit_log_df
    orig = orig_rows.iloc[0]

    edited_df.loc[i, 'edited_category']     = orig['raw_category']
    edited_df.loc[i, 'edited_raw_id']       = orig['raw_id']
    edited_df.loc[i, 'edited_source_file']  = orig['_source_file']
    edited_df.loc[i, 'edited_original_row'] = orig['_original_row']
    edited_df.loc[i, 'edited_pipeline_id']  = orig['pipeline_id']
    edited_df.loc[i, 'edited_point_no']     = orig['point_no']
    edited_df.loc[i, 'edited_point_no_int'] = orig['point_no_int']
    edited_df.loc[i, 'edited_x']            = orig['raw_x']
    edited_df.loc[i, 'edited_y']            = orig['raw_y']
    edited_df.loc[i, 'edited_z']            = orig['raw_z']
    edited_df.loc[i, 'edited_date']         = orig['raw_date']
    edited_df.loc[i, 'edited_x_val']        = orig['x_val']
    edited_df.loc[i, 'edited_y_val']        = orig['y_val']
    edited_df.loc[i, 'edited_z_val']        = orig['z_val']
    edited_df.loc[i, 'review_status']       = '未判讀'
    edited_df.loc[i, 'problem_type']        = '無'
    edited_df.loc[i, 'is_modified']         = False

    after_snap = _row_snapshot(edited_df.loc[i].to_dict(), _SNAP_FIELDS)
    orig_info  = {
        '_source_file':  curr.get('_source_file'),
        '_original_row': curr.get('_original_row'),
        'raw_id':        curr.get('raw_id'),
    }
    edit_log_df = _save_log(
        edit_log_df, point_key, '還原', orig_info,
        before_snap, after_snap,
        '未判讀', '無', curr.get('manual_note', ''),
        curr.get('include_in_result', '是'),
    )
    return edited_df, edit_log_df


def exclude_points(edited_df, edit_log_df, point_keys, note='手動排除'):
    """
    批次標記多個點位為「排除，不納入成果」。
    不刪除原始資料，只更新狀態，並寫入修正紀錄。
    """
    edited_df = edited_df.copy()
    for pkey in point_keys:
        idxs = edited_df.index[edited_df['point_key'] == pkey].tolist()
        if not idxs:
            continue
        i = idxs[0]
        prev = edited_df.loc[i].to_dict()
        before_snap = _row_snapshot(prev, _SNAP_FIELDS)

        edited_df.loc[i, 'review_status']     = '排除'
        edited_df.loc[i, 'include_in_result'] = '否'
        edited_df.loc[i, 'problem_type']      = '手動排除'
        old_note = str(edited_df.loc[i, 'manual_note'] or '')
        if note and note not in old_note:
            edited_df.loc[i, 'manual_note'] = (old_note + '；' + note).lstrip('；')
        edited_df.loc[i, 'is_modified'] = True

        after_snap = _row_snapshot(edited_df.loc[i].to_dict(), _SNAP_FIELDS)
        orig_info  = {
            '_source_file':  prev.get('_source_file'),
            '_original_row': prev.get('_original_row'),
            'raw_id':        prev.get('raw_id'),
        }
        edit_log_df = _save_log(
            edit_log_df, pkey, '手動排除', orig_info,
            before_snap, after_snap,
            '排除', '手動排除',
            str(edited_df.loc[i, 'manual_note']), '否',
        )
    return edited_df, edit_log_df


def restore_excluded(edited_df, edit_log_df, point_keys):
    """
    將指定的已排除點位還原為「未判讀，納入成果」。
    """
    edited_df = edited_df.copy()
    for pkey in point_keys:
        idxs = edited_df.index[edited_df['point_key'] == pkey].tolist()
        if not idxs:
            continue
        i = idxs[0]
        prev = edited_df.loc[i].to_dict()
        before_snap = _row_snapshot(prev, _SNAP_FIELDS)

        edited_df.loc[i, 'review_status']     = '未判讀'
        edited_df.loc[i, 'include_in_result'] = '是'
        edited_df.loc[i, 'problem_type']      = '無'
        edited_df.loc[i, 'is_modified']       = True

        after_snap = _row_snapshot(edited_df.loc[i].to_dict(), _SNAP_FIELDS)
        orig_info  = {
            '_source_file':  prev.get('_source_file'),
            '_original_row': prev.get('_original_row'),
            'raw_id':        prev.get('raw_id'),
        }
        edit_log_df = _save_log(
            edit_log_df, pkey, '還原排除', orig_info,
            before_snap, after_snap,
            '未判讀', '無', str(edited_df.loc[i, 'manual_note']), '是',
        )
    return edited_df, edit_log_df


def clear_all_edits(edited_df, original_df, edit_log_df):
    """還原所有人工修正至原始值，並寫入一筆清除紀錄。"""
    edited_df = edited_df.copy()
    n_mod = int(edited_df['is_modified'].sum())
    if n_mod == 0:
        return edited_df, edit_log_df

    for _, orig in original_df.iterrows():
        mask = (
            (edited_df['_source_file'] == orig['_source_file']) &
            (edited_df['_original_row'] == orig['_original_row'])
        )
        idxs = edited_df.index[mask].tolist()
        if not idxs:
            continue
        i = idxs[0]
        edited_df.loc[i, 'edited_category']     = orig['raw_category']
        edited_df.loc[i, 'edited_raw_id']       = orig['raw_id']
        edited_df.loc[i, 'edited_source_file']  = orig['_source_file']
        edited_df.loc[i, 'edited_original_row'] = orig['_original_row']
        edited_df.loc[i, 'edited_pipeline_id']  = orig['pipeline_id']
        edited_df.loc[i, 'edited_point_no']     = orig['point_no']
        edited_df.loc[i, 'edited_point_no_int'] = orig['point_no_int']
        edited_df.loc[i, 'edited_x']            = orig['raw_x']
        edited_df.loc[i, 'edited_y']            = orig['raw_y']
        edited_df.loc[i, 'edited_z']            = orig['raw_z']
        edited_df.loc[i, 'edited_date']         = orig['raw_date']
        edited_df.loc[i, 'edited_x_val']        = orig['x_val']
        edited_df.loc[i, 'edited_y_val']        = orig['y_val']
        edited_df.loc[i, 'edited_z_val']        = orig['z_val']
        edited_df.loc[i, 'review_status']       = '未判讀'
        edited_df.loc[i, 'problem_type']        = '無'
        edited_df.loc[i, 'is_modified']         = False

    log_row = {c: None for c in EDIT_LOG_COLS}
    log_row.update({
        '修正時間': _now_str(),
        '動作類型': '清除全部修正',
        '原始_來源CSV': f'（共清除 {n_mod} 筆修正）',
        '人工備註': f'清除全部：共 {n_mod} 筆人工修正已還原',
    })
    edit_log_df = pd.concat(
        [edit_log_df, pd.DataFrame([log_row])], ignore_index=True
    )
    return edited_df, edit_log_df


# ============================================================
# v3 點位畫布
# ============================================================

def _is_excluded(row):
    return (
        str(row.get('include_in_result', '是')) == '否'
        or str(row.get('review_status', '')) == '排除'
    )


def _is_pending(row):
    return str(row.get('review_status', '')) in ('待確認', '需廠商補測')


def _hover_v3(row):
    """產生 v3 版 hover 文字。"""
    status  = row.get('review_status', '')
    problem = row.get('problem_type', '')
    note    = row.get('manual_note', '')
    incl    = row.get('include_in_result', '是')
    is_mod  = bool(row.get('is_modified', False))

    parts = [
        f"<b>識別碼：</b>{row.get('edited_raw_id', row.get('raw_id', ''))}",
        f"<b>管線：</b>{row.get('edited_pipeline_id', '')}",
        f"<b>點號：</b>{row.get('edited_point_no', '')}",
        f"<b>X(E)：</b>{row.get('edited_x', '')}",
        f"<b>Y(N)：</b>{row.get('edited_y', '')}",
        f"<b>Z(H)：</b>{row.get('edited_z', '')}",
        f"<b>日期：</b>{row.get('edited_date', '')}",
        f"<b>類別碼：</b>{row.get('edited_category', row.get('raw_category', ''))}",
        '─────────────────',
        f"<b>判讀：</b>{status}　<b>問題：</b>{problem}",
    ]
    if note:
        parts.append(f"<b>備註：</b>{note}")
    parts.append(f"<b>納入成果：</b>{incl}")
    if is_mod:
        parts.append('⚙️ <b>已人工修正</b>')
    parts += [
        '─────────────────',
        f"<b>來源：</b>{row.get('_source_file', '')}",
        f"<b>列號：</b>{row.get('_original_row', '')}",
        '<i>點擊可在右側編輯</i>',
    ]
    return '<br>'.join(parts)


def build_figure_v3(edited_df, anomaly_set=None, selected_key=None,
                    canvas_height=580, uirevision='keep_zoom',
                    show_excluded=True):
    """
    以 edited_points_df 產生 v3 畫布。
    show_excluded=False 時，已排除點位（include_in_result=否 或 review_status=排除）不顯示於畫布。
    customdata 存 point_key，供 Streamlit plotly selection 使用。
    """
    if anomaly_set is None:
        anomaly_set = set()
    if not show_excluded:
        edited_df = edited_df[~edited_df.apply(_is_excluded, axis=1)].copy()

    df_plot = edited_df[
        edited_df['edited_x_val'].notna() & edited_df['edited_y_val'].notna()
    ].copy()

    if df_plot.empty:
        fig = go.Figure()
        fig.update_layout(
            title='GML 點位畫布（無有效座標）',
            height=canvas_height,
        )
        return fig

    traces = []
    pipeline_ids = sorted(df_plot['edited_pipeline_id'].dropna().unique())

    for i, pid in enumerate(pipeline_ids):
        color = COLOR_PALETTE[i % len(COLOR_PALETTE)]
        grp = df_plot[df_plot['edited_pipeline_id'] == pid].copy()

        def _skey(v):
            try:
                return (0, int(str(v).strip()))
            except Exception:
                return (1, 0)

        grp['_sk'] = grp['edited_point_no'].apply(_skey)
        grp = grp.sort_values('_sk')

        xs    = grp['edited_x_val'].tolist()
        ys    = grp['edited_y_val'].tolist()
        pkeys = grp['point_key'].tolist()

        # 連線 trace
        traces.append(go.Scatter(
            x=xs, y=ys,
            mode='lines',
            name=str(pid),
            line=dict(color=color, width=1.5),
            showlegend=False,
            hoverinfo='skip',
        ))

        # Per-point 樣式
        symbols, clrs, sizes, texts, border_colors, border_widths = [], [], [], [], [], []

        for _, row in grp.iterrows():
            pk      = row['point_key']
            excl    = _is_excluded(row)
            mod     = bool(row.get('is_modified', False))
            pend    = _is_pending(row)
            anom    = (row['_source_file'], row['_original_row']) in anomaly_set
            sel     = (pk == selected_key)

            texts.append(_hover_v3(row.to_dict()))

            if sel:
                symbols.append('circle-open')
                clrs.append('#FFD700')
                sizes.append(22)
                border_colors.append('#FFD700')
                border_widths.append(4)
            elif excl:
                symbols.append('x-thin')
                clrs.append('rgba(120,120,120,0.35)')
                sizes.append(9)
                border_colors.append('rgba(120,120,120,0.35)')
                border_widths.append(1.5)
            elif pend:
                symbols.append('triangle-up')
                clrs.append('#E67E22')
                sizes.append(11)
                border_colors.append('white')
                border_widths.append(1)
            elif mod:
                symbols.append('star')
                clrs.append('#E67E22')
                sizes.append(12)
                border_colors.append('white')
                border_widths.append(1)
            elif anom:
                symbols.append('x')
                clrs.append('red')
                sizes.append(11)
                border_colors.append('red')
                border_widths.append(2)
            else:
                symbols.append('circle')
                clrs.append(color)
                sizes.append(7)
                border_colors.append('white')
                border_widths.append(1)

        # point_no 標籤文字（用 text 參數顯示在畫布上）
        pno_labels = [str(row['edited_point_no'] or '') for _, row in grp.iterrows()]

        traces.append(go.Scatter(
            x=xs, y=ys,
            mode='markers+text',
            name=str(pid),
            marker=dict(
                symbol=symbols,
                color=clrs,
                size=sizes,
                line=dict(color=border_colors, width=border_widths),
            ),
            # text 用於畫布上的點號標籤
            text=pno_labels,
            textposition='top center',
            textfont=dict(size=8, color='#555'),
            # customdata 只存 point_key（字串），供點選事件取值
            customdata=pkeys,
            # hovertext 存 hover 內容，用 %{hovertext} 顯示
            hovertext=texts,
            hovertemplate='%{hovertext}<extra></extra>',
            legendgroup=str(pid),
            showlegend=True,
        ))

    # 無管線識別碼點位
    no_pid = df_plot[df_plot['edited_pipeline_id'].isna()]
    if not no_pid.empty:
        no_hover = [_hover_v3(r.to_dict()) for _, r in no_pid.iterrows()]
        traces.append(go.Scatter(
            x=no_pid['edited_x_val'].tolist(),
            y=no_pid['edited_y_val'].tolist(),
            mode='markers',
            name='（無管線識別碼）',
            marker=dict(color='gray', size=7, symbol='diamond'),
            customdata=no_pid['point_key'].tolist(),
            hovertext=no_hover,
            hovertemplate='%{hovertext}<extra></extra>',
        ))

    # 狀態圖例
    for sym, clr, lbl in [
        ('circle',        '#1f77b4',              '管線點位'),
        ('star',          '#E67E22',              '已修正 ⚙️'),
        ('triangle-up',   '#E67E22',              '待確認 ⚠️'),
        ('x',             'red',                  '異常 ✕'),
        ('x-thin',        'rgba(120,120,120,0.4)', '排除（灰）'),
        ('circle-open',   '#FFD700',              '選取中 ◎'),
    ]:
        traces.append(go.Scatter(
            x=[None], y=[None], mode='markers',
            marker=dict(symbol=sym, color=clr, size=9),
            name=lbl, showlegend=True,
            legendgroup='__status__',
        ))

    fig = go.Figure(data=traces)

    n_tot = len(df_plot)
    n_mod = int(edited_df['is_modified'].sum())
    n_ex  = int(edited_df.apply(_is_excluded, axis=1).sum())

    fig.update_layout(
        title=dict(text='GML 點位畫布 v3', font=dict(size=15)),
        xaxis=dict(
            title='X (E) 座標',
            scaleanchor='y', scaleratio=1,
            showgrid=True, gridcolor='#e0e0e0',
        ),
        yaxis=dict(title='Y (N) 座標', showgrid=True, gridcolor='#e0e0e0'),
        hovermode='closest',
        legend=dict(
            title='管線 / 狀態',
            bgcolor='rgba(255,255,255,0.88)',
            bordercolor='#ccc', borderwidth=1,
            tracegroupgap=4,
            font=dict(size=11),
        ),
        plot_bgcolor='#fafafa',
        paper_bgcolor='white',
        font=dict(family='Microsoft JhengHei, Arial, sans-serif', size=11),
        margin=dict(l=55, r=10, t=45, b=35),
        dragmode='pan',
        height=canvas_height,
        uirevision=uirevision,
    )
    fig.add_annotation(
        text=f'共 {n_tot} 點　已修正 {n_mod}　已排除 {n_ex}',
        xref='paper', yref='paper', x=0, y=-0.06,
        showarrow=False, font=dict(size=10, color='#666'), align='left',
    )
    return fig


# ============================================================
# 從事件取 point_key
# ============================================================

def extract_point_key(event):
    """從 st.plotly_chart on_select 事件安全取出 point_key。"""
    try:
        if event is None:
            return None
        sel = getattr(event, 'selection', None)
        if sel is None:
            return None
        pts = getattr(sel, 'points', None) or []
        if not pts:
            return None
        pt = pts[0]
        cd = pt.get('customdata') if isinstance(pt, dict) else getattr(pt, 'customdata', None)
        if cd is None:
            return None
        if isinstance(cd, (list, tuple)):
            return str(cd[0]) if cd else None
        return str(cd)
    except Exception:
        return None


# ============================================================
# 多工作表 Excel 匯出（v3）
# ============================================================

_THIN   = Side(style='thin', color='CCCCCC')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HDR_FONT  = Font(name='微軟正黑體', bold=True, color='FFFFFF', size=10)
_BODY_FONT = Font(name='微軟正黑體', size=10)
_CENTER = Alignment(horizontal='center', vertical='center')
_LEFT   = Alignment(horizontal='left',   vertical='center')

_FILLS = {
    'blue':   PatternFill('solid', fgColor='1F4E79'),
    'green':  PatternFill('solid', fgColor='1A5632'),
    'orange': PatternFill('solid', fgColor='7E5109'),
    'red':    PatternFill('solid', fgColor='922B21'),
    'purple': PatternFill('solid', fgColor='4A235A'),
    'teal':   PatternFill('solid', fgColor='0E6251'),
}
_ALT_FILL  = PatternFill('solid', fgColor='EBF5FB')
_MOD_FILL  = PatternFill('solid', fgColor='FEF9E7')
_EXCL_FILL = PatternFill('solid', fgColor='F2F3F4')


def _write_sheet(ws, headers, rows, fill_key='blue'):
    fill = _FILLS[fill_key]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = _HDR_FONT
        cell.fill = fill
        cell.alignment = _CENTER
        cell.border = _BORDER
    ws.row_dimensions[1].height = 20

    for r, row_data in enumerate(rows, 2):
        is_alt  = (r % 2 == 0)
        is_mod  = bool(row_data.get('__is_modified__', False)) if isinstance(row_data, dict) else False
        is_excl = bool(row_data.get('__is_excluded__', False)) if isinstance(row_data, dict) else False

        for c, h in enumerate(headers, 1):
            if isinstance(row_data, dict):
                val = row_data.get(h, '')
            else:
                val = row_data[c - 1] if c - 1 < len(row_data) else ''
            if val is None or (isinstance(val, float) and math.isnan(val)):
                val = ''
            cell = ws.cell(row=r, column=c, value=val)
            cell.font      = _BODY_FONT
            cell.border    = _BORDER
            cell.alignment = _LEFT
            if is_excl:
                cell.fill = _EXCL_FILL
            elif is_mod:
                cell.fill = _MOD_FILL
            elif is_alt:
                cell.fill = _ALT_FILL

    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value else 0 for cell in col),
            default=4,
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, 8), 45)

    ws.freeze_panes = 'A2'


_CLS_SHEET_FILLS = ['blue', 'green', 'teal', 'orange', 'red', 'purple', 'teal', 'green', 'blue']


def export_classification_bytes(edited_df, facility_type='全部設施'):
    """
    匯出分類表 Excel BytesIO。
    facility_type='全部設施' → 輸出 9 張工作表（每種設施一張）。
    其他 facility_type → 輸出單一設施類型（一張工作表）。
    預設只含納入成果的點位（已排除不輸出）。
    """
    wb = openpyxl.Workbook()
    if facility_type == '全部設施':
        for idx, ft in enumerate(FACILITY_TYPES):
            ws = wb.active if idx == 0 else wb.create_sheet()
            ws.title = f'分類表_{ft}'
            cls_df = build_classification_df(edited_df, ft, include_excluded=False)
            _write_sheet(ws, list(cls_df.columns), cls_df.to_dict('records'),
                         _CLS_SHEET_FILLS[idx])
    else:
        idx = FACILITY_TYPES.index(facility_type) if facility_type in FACILITY_TYPES else 0
        ws = wb.active
        ws.title = f'分類表_{facility_type}'
        cls_df = build_classification_df(edited_df, facility_type, include_excluded=False)
        _write_sheet(ws, list(cls_df.columns), cls_df.to_dict('records'),
                     _CLS_SHEET_FILLS[idx])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_v3_bytes(original_df, edited_df, edit_log_df, anomaly_list, dup_list):
    """
    產生 v3 多工作表 Excel BytesIO（16 張工作表）：
      - 6 基礎表（原始解析 / 修正後成果 / 人工判讀 / 修正紀錄 / 異常清單 / 重複量測）
      - 9 分類表（管線 ～ 場站）
      - 1 已排除資料表
    修正後成果 / 人工判讀 / 分類表：預設只含 include_in_result='是' 的點位。
    """
    wb = openpyxl.Workbook()

    # 區分納入成果 vs 已排除
    included_df = edited_df[~edited_df.apply(_is_excluded, axis=1)]
    excluded_df = edited_df[edited_df.apply(_is_excluded, axis=1)]

    # ── 1. 原始解析點位表（永遠保留全部）─────────────
    ws1 = wb.active
    ws1.title = '原始解析點位表'
    orig_h = ['來源CSV', '原始列號', '原始識別碼', '管線識別碼', '點號', '類別碼',
              'X', 'Y', 'Z', '測量日期']
    orig_m = {
        '來源CSV': '_source_file', '原始列號': '_original_row',
        '原始識別碼': 'raw_id', '管線識別碼': 'pipeline_id',
        '點號': 'point_no', '類別碼': 'raw_category',
        'X': 'raw_x', 'Y': 'raw_y', 'Z': 'raw_z', '測量日期': 'raw_date',
    }
    _write_sheet(ws1, orig_h, [{h: r.get(orig_m[h]) for h in orig_h} for _, r in original_df.iterrows()], 'blue')

    # ── 2. 修正後點位表（僅含納入成果的點位）──────────
    ws2 = wb.create_sheet('修正後點位表')
    edit_h = [
        '來源CSV', '原始列號', '原始識別碼', '類別碼',
        '管線識別碼', '點號', 'X', 'Y', 'Z', '測量日期',
        '判讀狀態', '問題類型', '人工備註', '是否納入成果', '是否已修正',
    ]
    edit_m = {
        '來源CSV': 'edited_source_file', '原始列號': 'edited_original_row',
        '原始識別碼': 'edited_raw_id', '類別碼': 'edited_category',
        '管線識別碼': 'edited_pipeline_id', '點號': 'edited_point_no',
        'X': 'edited_x', 'Y': 'edited_y', 'Z': 'edited_z', '測量日期': 'edited_date',
        '判讀狀態': 'review_status', '問題類型': 'problem_type',
        '人工備註': 'manual_note', '是否納入成果': 'include_in_result',
        '是否已修正': 'is_modified',
    }
    edit_rows = []
    for _, row in included_df.iterrows():
        d = {h: row.get(edit_m[h]) for h in edit_h}
        d['是否已修正'] = '是' if d.get('是否已修正') else '否'
        d['__is_modified__'] = bool(row.get('is_modified', False))
        d['__is_excluded__'] = False
        edit_rows.append(d)
    _write_sheet(ws2, edit_h, edit_rows, 'green')

    # ── 3. 人工判讀紀錄表（僅含納入成果的點位）──────
    ws3 = wb.create_sheet('人工判讀紀錄表')
    rev_h = [
        '來源CSV', '原始列號', '原始識別碼', '管線識別碼', '點號',
        'X', 'Y', 'Z', '測量日期',
        '判讀狀態', '問題類型', '人工備註', '是否納入成果',
    ]
    rev_m = {
        '來源CSV': '_source_file', '原始列號': '_original_row',
        '原始識別碼': 'raw_id', '管線識別碼': 'edited_pipeline_id',
        '點號': 'edited_point_no',
        'X': 'edited_x', 'Y': 'edited_y', 'Z': 'edited_z',
        '測量日期': 'edited_date',
        '判讀狀態': 'review_status', '問題類型': 'problem_type',
        '人工備註': 'manual_note', '是否納入成果': 'include_in_result',
    }
    _write_sheet(ws3, rev_h, [{h: r.get(rev_m[h]) for h in rev_h} for _, r in included_df.iterrows()], 'teal')

    # ── 4. 修正紀錄表 ────────────────────────────────
    ws4 = wb.create_sheet('修正紀錄表')
    log_rows = edit_log_df.to_dict('records') if not edit_log_df.empty else []
    _write_sheet(ws4, EDIT_LOG_COLS, log_rows, 'orange')

    # ── 5. 異常清單 ──────────────────────────────────
    ws5 = wb.create_sheet('異常清單')
    anom_h = [
        '異常類型', '異常說明', '原始識別碼', '管線識別碼', '點號',
        'X', 'Y', 'Z', '測量日期', '來源CSV', '原始列號', '建議處理方式',
    ]
    _write_sheet(ws5, anom_h, anomaly_list, 'red')

    # ── 6. 重複量測比對表 ────────────────────────────
    ws6 = wb.create_sheet('重複量測比對表')
    dup_h = [
        '來源CSV', '原始列號', '原始識別碼', '管線識別碼', '點號',
        '測量日期', 'X', 'Y', 'Z',
        '比對基準來源CSV', '比對基準原始列號',
        '與比對點XY差距(m)', '與比對點Z差距(m)',
        '判定結果', '建議處理方式',
    ]
    _write_sheet(ws6, dup_h, dup_list, 'purple')

    # ── 7-15. 分類表（表 6-9 ～ 表 6-17，僅含納入成果的點位）──
    for idx, ft in enumerate(FACILITY_TYPES):
        ws_cls = wb.create_sheet(f'分類表_{ft}')
        cls_df = build_classification_df(edited_df, ft, include_excluded=False)
        _write_sheet(ws_cls, list(cls_df.columns), cls_df.to_dict('records'),
                     _CLS_SHEET_FILLS[idx])

    # ── 16. 已排除資料 ────────────────────────────────
    ws_excl = wb.create_sheet('已排除資料')
    excl_h = [
        'point_key', '來源CSV', '原始列號', '原始識別碼',
        '類別碼', '管線識別碼', '點號', 'X', 'Y', 'Z',
        '判讀狀態', '問題類型', '人工備註', '是否納入成果',
    ]
    excl_m = {
        'point_key': 'point_key',
        '來源CSV': '_source_file', '原始列號': '_original_row',
        '原始識別碼': 'edited_raw_id', '類別碼': 'edited_category',
        '管線識別碼': 'edited_pipeline_id', '點號': 'edited_point_no',
        'X': 'edited_x', 'Y': 'edited_y', 'Z': 'edited_z',
        '判讀狀態': 'review_status', '問題類型': 'problem_type',
        '人工備註': 'manual_note', '是否納入成果': 'include_in_result',
    }
    excl_rows = [{h: r.get(excl_m[h]) for h in excl_h} for _, r in excluded_df.iterrows()]
    _write_sheet(ws_excl, excl_h, excl_rows, 'red')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
