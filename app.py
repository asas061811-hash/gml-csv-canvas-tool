"""
app.py - GML 點位處理工具 V0.6 版

佈局：
  左側 sidebar ─ 功能模式切換（資料整理 / GML 輸出）+ 模式專屬控制
  主畫面依模式切換頁籤：
    資料整理：畫布 + 編輯 + 修正後點位表 / 原始解析 / 異常 / 重複 / 修正紀錄 / 分類表 / 下載
    GML 輸出：GML 產製 / GML 預覽 / GML 下載 / GML 檢核訊息
"""

import hashlib
import io
import math
import os
import sys
import time

import openpyxl
import pandas as pd
import plotly.io as pio
import streamlit as st

from src.anomaly_detector import detect_anomalies
from src.canvas_generator import build_figure
from src.csv_loader import load_uploaded_csvs
from src.excel_exporter import (
    CORRECTION_HEADERS,
    _export_to_bytes,
    export_anomalies,
    export_correction_template,
    export_duplicates,
    export_parsed_points,
    export_review_template,
)
from src.id_parser import parse_all_ids
from src.v3_helpers import (
    FACILITY_SCHEMAS,
    FACILITY_TYPES,
    INCLUDE_RESULT_OPTIONS,
    PROBLEM_TYPE_OPTIONS,
    REVIEW_STATUS_OPTIONS,
    build_classification_df,
    build_cls_pkey_list,
    build_figure_v3,
    clear_all_edits,
    count_pipelines,
    decode_category,
    exclude_points,
    export_classification_bytes,
    export_v3_bytes,
    extract_point_key,
    get_facility_type,
    init_edit_log,
    init_edited_points_df,
    pipeline_export_rename,
    quick_mark,
    restore_excluded,
    restore_point,
    reverse_pipeline_order,
    save_point_edit,
)
from src.attribute_merger import (
    load_code_mappings,
    load_definition_excel,
    merge_definition_attributes,
    convert_definition_to_code,
    convert_code_to_definition,
    build_code_definition_table,
    validate_merge_result,
)
from src.gml_generator import (
    TARGET_OPTIONS,
    TARGET_TAOYUAN,
    ELEVATION_OPTIONS,
    ELEVATION_ABSOLUTE,
    generate_gml,
    is_implemented,
    export_gml_wide_csv_bytes,
)

CONFIG_PATH = os.path.join(os.path.dirname(__file__), 'config', 'rules.yaml')
ACTION_TYPE_OPTIONS = ['修正', '採用', '排除', '新增', '刪除']

DEFAULT_XY_TOL = 0.30
DEFAULT_Z_TOL  = 0.10

_MIME_XLSX = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


# ============================================================
# 工具函式
# ============================================================

def build_anomaly_set(anomaly_list):
    s = set()
    for item in anomaly_list:
        src    = item.get('來源CSV')
        row_no = item.get('原始列號')
        if src and row_no and str(row_no) != '（整份檔案）':
            try:
                s.add((src, int(row_no)))
            except (ValueError, TypeError):
                pass
    return s


def _sk(selected_key):
    return hashlib.md5((selected_key or '').encode('utf-8')).hexdigest()[:12]


def _v(val):
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return ''
    return str(val)


def _is_excl(row):
    return (str(row.get('include_in_result', '是')) == '否' or
            str(row.get('review_status', '')) == '排除')


def make_anomaly_styler(anom_df):
    HIGH = {'座標空值', 'X 座標格式異常', 'Y 座標格式異常', '重複點位座標差異過大'}
    def color_row(row):
        return (['background-color: #fdecea'] * len(row)
                if row.get('異常類型') in HIGH else [''] * len(row))
    return anom_df.style.apply(color_row, axis=1)


# ============================================================
# 右側大型編輯視窗
# ============================================================

def render_edit_panel(edited_df, original_df, anomaly_set):
    selected_key = st.session_state.get('selected_key')
    st.markdown('## 📋 點位資料編輯')

    if not selected_key:
        st.info(
            '**請在左側畫布上點選一個點位**\n\n'
            '點選後，此處將顯示該點位的所有欄位，\n'
            '您可以逐一修改並儲存。',
            icon='👈',
        )
        return

    if selected_key not in edited_df['point_key'].values:
        st.warning('找不到選取的點位，請重新點選。')
        if st.button('清除選取', key='clear_sel_err'):
            st.session_state.selected_key = None
            st.rerun()
        return

    row = edited_df[edited_df['point_key'] == selected_key].iloc[0]
    k   = _sk(selected_key)

    is_anom = (row['_source_file'], row['_original_row']) in anomaly_set
    is_mod  = bool(row.get('is_modified', False))
    status  = str(row.get('review_status', ''))

    badges = []
    if is_anom:  badges.append('⚠️ 異常點位')
    if is_mod:   badges.append('⚙️ 已修正')
    if status == '排除':    badges.append('🚫 已排除')
    elif status == '正常':  badges.append('✅ 正常')
    elif status == '待確認': badges.append('❓ 待確認')
    if badges:
        st.markdown('　'.join(f'`{b}`' for b in badges))

    qc1, qc2, qc3 = st.columns(3)
    with qc1:
        if st.button('✅ 標記正常', key=f'qnorm_{k}', use_container_width=True):
            new_edf, new_log = quick_mark(
                st.session_state.edited_points_df,
                st.session_state.edit_log_df, selected_key, '標記正常')
            st.session_state.edited_points_df = new_edf
            st.session_state.edit_log_df = new_log
            st.rerun()
    with qc2:
        if st.button('🚫 標記排除', key=f'qexcl_{k}', use_container_width=True):
            new_edf, new_log = quick_mark(
                st.session_state.edited_points_df,
                st.session_state.edit_log_df, selected_key, '標記排除')
            st.session_state.edited_points_df = new_edf
            st.session_state.edit_log_df = new_log
            st.rerun()
    with qc3:
        if st.button('↩️ 還原此點', key=f'restore_{k}',
                     disabled=not is_mod, use_container_width=True):
            new_edf, new_log = restore_point(
                st.session_state.edited_points_df,
                st.session_state.edit_log_df, selected_key, original_df)
            st.session_state.edited_points_df = new_edf
            st.session_state.edit_log_df = new_log
            st.rerun()

    # ── 點位排序轉向 ─────────────────────────────
    cur_pid = _v(row.get('edited_pipeline_id'))
    if cur_pid:
        with st.expander(f'🔄 點位排序轉向（管線：{cur_pid}）', expanded=False):
            pipe_mask = edited_df['edited_pipeline_id'].astype(str).str.strip() == cur_pid
            pipe_count = int(pipe_mask.sum())
            st.caption(f'此管線共 {pipe_count} 個點位。反轉後點號將重新從 1 開始編號。')
            if pipe_count <= 1:
                st.info('此管線只有 1 個點，無需排序轉向。')
            else:
                if not st.session_state.get(f'confirm_reverse_{k}', False):
                    if st.button('🔄 反轉此管線點位順序', key=f'rev_btn_{k}',
                                 use_container_width=True):
                        st.session_state[f'confirm_reverse_{k}'] = True
                        st.rerun()
                else:
                    st.warning('確定要反轉此管線點位順序嗎？此操作會更新點號、畫布、分類表與 GML 輸出座標順序。')
                    rc1, rc2 = st.columns(2)
                    if rc1.button('✅ 確認反轉', key=f'rev_yes_{k}', type='primary',
                                  use_container_width=True):
                        new_edf, new_log, msg = reverse_pipeline_order(
                            st.session_state.edited_points_df,
                            st.session_state.edit_log_df, cur_pid)
                        st.session_state.edited_points_df = new_edf
                        st.session_state.edit_log_df = new_log
                        st.session_state[f'confirm_reverse_{k}'] = False
                        st.toast(msg, icon='🔄')
                        st.rerun()
                    if rc2.button('❌ 取消', key=f'rev_no_{k}', use_container_width=True):
                        st.session_state[f'confirm_reverse_{k}'] = False
                        st.rerun()

    st.divider()

    with st.form(key=f'edit_form_{k}'):
        st.markdown('#### 識別與分類')
        fa, fb = st.columns(2)
        new_rid = fa.text_input('原始識別碼',
            value=_v(row.get('edited_raw_id', row.get('raw_id'))), key=f'v3_rid_{k}',
            help='可編輯；系統另保留原始值供追溯')
        new_cat = fb.text_input('類別碼',
            value=_v(row.get('edited_category', row.get('raw_category'))), key=f'v3_cat_{k}')
        _cat_val = _v(row.get('edited_category', row.get('raw_category')))
        _cat_info = decode_category(_cat_val)
        if _cat_val and _cat_info['parse_ok']:
            st.caption(
                f"📋 **{_cat_info['major_name']}** ｜ "
                f"中類 {_cat_info['mid_code']} {_cat_info['mid_name']} ｜ "
                f"小類 {_cat_info['minor_code']} {_cat_info['minor_name']} ｜ "
                f"細類 {_cat_info['fine_code']} {_cat_info['fine_name']}")
        elif _cat_val:
            st.caption('⚠️ 類別碼無法解析（非標準 7 碼數字格式）')
        fc, fd = st.columns(2)
        new_pid = fc.text_input('管線識別碼',
            value=_v(row.get('edited_pipeline_id')), key=f'v3_pid_{k}')
        new_pno = fd.text_input('點號',
            value=_v(row.get('edited_point_no')), key=f'v3_pno_{k}')

        st.markdown('#### 座標')
        cx, cy, cz = st.columns(3)
        new_x = cx.text_input('X / E', value=_v(row.get('edited_x')), key=f'v3_x_{k}',
            help='修改後畫布點位位置將立即更新')
        new_y = cy.text_input('Y / N', value=_v(row.get('edited_y')), key=f'v3_y_{k}',
            help='修改後畫布點位位置將立即更新')
        new_z = cz.text_input('Z / H', value=_v(row.get('edited_z')), key=f'v3_z_{k}')

        st.markdown('#### 其他屬性')
        new_date = st.text_input('測量日期',
            value=_v(row.get('edited_date')), key=f'v3_date_{k}', placeholder='YYYY-MM-DD 或原始格式')
        ge, gf = st.columns(2)
        new_src = ge.text_input('來源 CSV（可修改顯示值）',
            value=_v(row.get('edited_source_file', row.get('_source_file'))), key=f'v3_src_{k}',
            help='系統另保留原始追溯值')
        new_rrow = gf.text_input('原始列號（可修改顯示值）',
            value=_v(row.get('edited_original_row', row.get('_original_row'))), key=f'v3_rrow_{k}',
            help='系統另保留原始追溯值')

        st.divider()
        st.markdown('#### 判讀與標記')
        cur_status = row.get('review_status', '未判讀')
        si = REVIEW_STATUS_OPTIONS.index(cur_status) if cur_status in REVIEW_STATUS_OPTIONS else 0
        new_status = st.selectbox('判讀狀態', REVIEW_STATUS_OPTIONS, index=si, key=f'v3_status_{k}')
        cur_prob = row.get('problem_type', '無')
        pi = PROBLEM_TYPE_OPTIONS.index(cur_prob) if cur_prob in PROBLEM_TYPE_OPTIONS else 0
        new_prob = st.selectbox('問題類型', PROBLEM_TYPE_OPTIONS, index=pi, key=f'v3_prob_{k}')
        cur_incl = str(row.get('include_in_result', '是'))
        ii = INCLUDE_RESULT_OPTIONS.index(cur_incl) if cur_incl in INCLUDE_RESULT_OPTIONS else 0
        new_incl = st.selectbox('是否納入成果', INCLUDE_RESULT_OPTIONS, index=ii, key=f'v3_incl_{k}')
        new_note = st.text_area('人工備註', value=_v(row.get('manual_note')), height=90, key=f'v3_note_{k}')
        new_action = st.selectbox('動作類型（記錄用）', ACTION_TYPE_OPTIONS, index=0, key=f'v3_action_{k}')

        st.divider()
        save_btn = st.form_submit_button('💾 儲存修改', type='primary', use_container_width=True)

    if save_btn:
        form_data = {
            'raw_id': new_rid, 'category': new_cat, 'pipeline_id': new_pid,
            'point_no': new_pno, 'x': new_x, 'y': new_y, 'z': new_z,
            'date': new_date, 'source_file': new_src, 'original_row': new_rrow,
            'review_status': new_status, 'problem_type': new_prob,
            'manual_note': new_note, 'include_in_result': new_incl,
            'action_type': new_action,
        }
        new_edf, new_log = save_point_edit(
            st.session_state.edited_points_df,
            st.session_state.edit_log_df, selected_key, form_data)
        st.session_state.edited_points_df = new_edf
        st.session_state.edit_log_df = new_log
        st.toast('✅ 已儲存修改！', icon='✅')
        st.rerun()

    st.divider()
    with st.expander('🔒 原始追溯資料（不可更改，供修正紀錄使用）', expanded=False):
        oa, ob = st.columns(2)
        oa.text_input('原始來源 CSV', value=_v(row.get('_source_file')), disabled=True, key=f'orig_src_{k}')
        ob.text_input('原始列號', value=_v(row.get('_original_row')), disabled=True, key=f'orig_row_{k}')
        oc, od = st.columns(2)
        oc.text_input('原始識別碼', value=_v(row.get('raw_id')), disabled=True, key=f'orig_rid_{k}')
        od.text_input('原始管線識別碼', value=_v(row.get('pipeline_id')), disabled=True, key=f'orig_pid_{k}')
        oe, of_, og = st.columns(3)
        oe.text_input('原始 X', value=_v(row.get('raw_x')), disabled=True, key=f'orig_x_{k}')
        of_.text_input('原始 Y', value=_v(row.get('raw_y')), disabled=True, key=f'orig_y_{k}')
        og.text_input('原始 Z', value=_v(row.get('raw_z')), disabled=True, key=f'orig_z_{k}')
        st.text_input('原始測量日期', value=_v(row.get('raw_date')), disabled=True, key=f'orig_date_{k}')
        st.text_input('point_key', value=str(selected_key), disabled=True, key=f'orig_pk_{k}')

    n_mod_total = int(st.session_state.edited_points_df['is_modified'].sum())
    if n_mod_total > 0:
        st.divider()
        if not st.session_state.get('confirm_clear', False):
            if st.button(f'🗑️ 清除全部修正（{n_mod_total} 筆）', type='secondary', use_container_width=True):
                st.session_state.confirm_clear = True
                st.rerun()
        else:
            st.error(f'確定清除全部 {n_mod_total} 筆修正？此動作不可還原（會寫入紀錄）。')
            cc1, cc2 = st.columns(2)
            with cc1:
                if st.button('✅ 確認清除', type='primary', use_container_width=True):
                    new_edf, new_log = clear_all_edits(
                        st.session_state.edited_points_df, original_df,
                        st.session_state.edit_log_df)
                    st.session_state.edited_points_df = new_edf
                    st.session_state.edit_log_df = new_log
                    st.session_state.confirm_clear = False
                    st.rerun()
            with cc2:
                if st.button('❌ 取消', use_container_width=True):
                    st.session_state.confirm_clear = False
                    st.rerun()


# ============================================================
# 主程式
# ============================================================

def main():
    st.set_page_config(
        page_title='GML 點位處理工具 V0.6 版',
        page_icon='📍',
        layout='wide',
        initial_sidebar_state='expanded',
    )

    # ── Session State 初始化 ─────────────────────────
    for key, default in [
        ('processed',         False),
        ('df',                None),
        ('file_meta',         None),
        ('anomaly_list',      []),
        ('duplicate_list',    []),
        ('anomaly_set',       set()),
        ('edited_points_df',  None),
        ('edit_log_df',       None),
        ('selected_key',      None),
        ('load_errors',       []),
        ('confirm_clear',     False),
        ('xy_tol_used',       DEFAULT_XY_TOL),
        ('z_tol_used',        DEFAULT_Z_TOL),
        ('canvas_revision',   'init'),
        ('show_excluded',     False),
        ('confirm_excl_tab0', False),
        ('pending_excl_tab0', []),
        ('confirm_excl_cls',  False),
        ('pending_excl_cls',  []),
        ('app_mode',          '資料整理'),
        ('gml_result_text',   ''),
        ('gml_result_ok',     0),
        ('gml_result_errors', []),
        ('gml_result_facility', ''),
        ('gml_result_target', ''),
        ('attr_merged', {}),
        ('attr_code_dfs', {}),
    ]:
        if key not in st.session_state:
            st.session_state[key] = default

    # ════════════════════════════════════════════════
    # 左側 Sidebar
    # ════════════════════════════════════════════════
    with st.sidebar:
        st.title('📍 GML 點位處理工具')
        st.caption('V0.6 版')
        st.divider()

        # ── 功能模式切換 ───────────────────────────
        app_mode = st.segmented_control(
            '功能模式',
            ['資料整理', 'GML 輸出'],
            default='資料整理',
            key='app_mode',
        )
        if app_mode is None:
            app_mode = '資料整理'

        st.divider()

        if app_mode == '資料整理':
            # ① 上傳 CSV
            st.header('① 上傳 CSV')
            uploaded = st.file_uploader(
                '選擇一或多個 CSV 檔案',
                type=['csv', 'CSV'],
                accept_multiple_files=True,
                help='支援 UTF-8、UTF-8 BOM、Big5、CP950 等編碼',
            )
            if uploaded:
                st.success(f'已選取 {len(uploaded)} 個檔案')
                for f in uploaded:
                    st.caption(f'• {f.name}')
            st.divider()

            # ② 重複點位判定門檻
            st.header('② 重複點位判定門檻')
            xy_tol = st.slider('XY 距離門檻（公尺）',
                min_value=0.01, max_value=2.00, value=DEFAULT_XY_TOL, step=0.01,
                help='兩重複點的平面距離超過此值 → 標記「差異過大」')
            z_tol = st.slider('Z 高程門檻（公尺）',
                min_value=0.01, max_value=1.00, value=DEFAULT_Z_TOL, step=0.01,
                help='兩重複點的高程差超過此值 → 標記「差異過大」')
            st.caption(f'目前設定：XY ≤ {xy_tol:.2f}m　Z ≤ {z_tol:.2f}m')
            if st.session_state.processed:
                st.caption(
                    f'上次處理使用：XY={st.session_state.xy_tol_used:.2f}m　'
                    f'Z={st.session_state.z_tol_used:.2f}m')
            st.divider()

            # ③ 開始處理
            st.header('③ 執行處理')
            run_btn = st.button('▶ 開始處理', type='primary',
                disabled=not bool(uploaded), use_container_width=True,
                help='處理後所有修正記錄將被清除')

            # ④ 處理結果摘要
            if st.session_state.processed:
                st.divider()
                st.header('處理結果摘要')
                edf = st.session_state.edited_points_df
                al  = st.session_state.anomaly_list
                dl  = st.session_state.duplicate_list
                c1, c2 = st.columns(2)
                c1.metric('點位總數', len(edf))
                c2.metric('管線數', count_pipelines(edf))
                c1.metric('異常筆數', len(al))
                c2.metric('重複記錄', len(dl))
                n_excl_sb = int(edf.apply(_is_excl, axis=1).sum())
                st.caption(f'管線數依「管線識別碼」唯一值統計　已排除 {n_excl_sb} 筆')
                n_mod = int(edf['is_modified'].sum())
                if n_mod:
                    st.info(f'⚙️ 已修正 {n_mod} 個點位')
                st.divider()

                # ④ 顯示排除資料
                st.session_state.show_excluded = st.toggle(
                    '顯示排除資料', value=st.session_state.get('show_excluded', False),
                    help='開啟後，畫布與表格將顯示已排除的點位（灰色）',
                    key='sidebar_show_excl')
                if st.session_state.load_errors:
                    st.warning(f'{len(st.session_state.load_errors)} 個檔案讀取失敗')

        else:  # GML 輸出模式
            uploaded = None
            run_btn = False
            xy_tol = DEFAULT_XY_TOL
            z_tol = DEFAULT_Z_TOL

            # ① 資料來源
            st.header('① 資料來源')
            gml_source = st.radio('資料來源',
                ['使用目前修正後資料', '上傳 GML 專用寬表 CSV'],
                key='gml_source_sb')
            st.divider()

            # ② 目標規範
            st.header('② 目標規範')
            target_keys = list(TARGET_OPTIONS.keys())
            target_labels = list(TARGET_OPTIONS.values())
            gml_target_idx = st.selectbox('目標規範',
                range(len(target_keys)),
                format_func=lambda i: target_labels[i],
                key='gml_target_sb')
            gml_target = target_keys[gml_target_idx]
            st.divider()

            # ③ 設施類型
            st.header('③ 設施類型')
            gml_facility = st.radio('設施類型', FACILITY_TYPES, key='gml_facility_sb')
            st.divider()

            # ④ Z 值高程模式
            st.header('④ Z 值高程模式')
            elev_keys = list(ELEVATION_OPTIONS.keys())
            elev_labels = list(ELEVATION_OPTIONS.values())
            gml_elev_idx = st.selectbox('高程模式',
                range(len(elev_keys)),
                format_func=lambda i: elev_labels[i],
                key='gml_elev_sb')
            gml_elevation = elev_keys[gml_elev_idx]
            st.divider()

            # ⑤ GML 產出
            st.header('⑤ GML 產出')
            has_data = st.session_state.processed or gml_source == '上傳 GML 專用寬表 CSV'
            if not has_data:
                st.info('請先完成資料整理，或上傳 GML 專用寬表 CSV')

    # ════════════════════════════════════════════════
    # 處理邏輯（資料整理模式）
    # ════════════════════════════════════════════════
    if app_mode == '資料整理' and run_btn and uploaded:
        progress = st.progress(0, '讀取 CSV...')
        df, file_meta, errors = load_uploaded_csvs(uploaded, CONFIG_PATH)
        progress.progress(20, '解析識別碼...')

        if df.empty:
            st.error('未讀取到任何資料，請確認 CSV 格式是否正確。')
            st.stop()

        df = parse_all_ids(df)
        progress.progress(45, f'異常偵測（XY≤{xy_tol}m, Z≤{z_tol}m）...')

        df, anomaly_list, dup_list = detect_anomalies(
            df, CONFIG_PATH, file_meta,
            xy_tolerance=xy_tol, z_tolerance=z_tol)
        progress.progress(70, '初始化可編輯點位表...')

        anomaly_set      = build_anomaly_set(anomaly_list)
        edited_points_df = init_edited_points_df(df)
        edit_log_df      = init_edit_log()
        progress.progress(100, '完成！')
        progress.empty()

        st.session_state.update({
            'df': df, 'file_meta': file_meta,
            'anomaly_list': anomaly_list, 'duplicate_list': dup_list,
            'anomaly_set': anomaly_set,
            'edited_points_df': edited_points_df, 'edit_log_df': edit_log_df,
            'selected_key': None, 'load_errors': errors,
            'processed': True, 'confirm_clear': False,
            'xy_tol_used': xy_tol, 'z_tol_used': z_tol,
            'canvas_revision': f'rev_{int(time.time())}',
        })

    # ════════════════════════════════════════════════
    # 歡迎畫面（尚未處理且為資料整理模式）
    # ════════════════════════════════════════════════
    if not st.session_state.processed and app_mode == '資料整理':
        st.markdown("""
## 歡迎使用 GML 點位處理工具 V0.6 版

### 第一階段：資料整理
1. 在左側 **① 上傳 CSV** — 可多選
2. 調整 **② 重複點位判定門檻**
3. 按 **▶ 開始處理**
4. 在畫布**點選點位** → 右側**編輯視窗**載入資料
5. 修改欄位 → **儲存修改**
6. 查看分類表、下載 Excel 成果

### 第二階段：GML 輸出
1. 左側切換至 **GML 輸出** 模式
2. 選擇資料來源、目標規範、設施類型
3. 產生 GML → 預覽 → 下載

---
#### 畫布符號說明
| 符號 | 意義 |
|------|------|
| ● 管線色圓形 | 一般點位 |
| ⭐ 橘色星星 | 已人工修正 |
| ▲ 橘色三角 | 待確認 / 需廠商補測 |
| ✕ 紅色 | 異常點位 |
| ✕ 灰色 | 已排除 |
| ○ 金色大圓 | 目前選取 |

> 原始 CSV 永遠不會被修改。所有修正均寫入修正紀錄表，可完整追溯。
        """)
        return

    if not st.session_state.processed and app_mode == 'GML 輸出':
        st.markdown("""
## GML 輸出模式

請先完成資料整理（切換至「資料整理」模式並上傳 CSV），或在左側選擇「上傳 GML 專用寬表 CSV」作為資料來源。
        """)
        # 允許直接上傳寬表
        if st.session_state.get('gml_source_sb') == '上傳 GML 專用寬表 CSV':
            _render_gml_mode_tabs()
        return

    # ════════════════════════════════════════════════
    # 取出 session 資料
    # ════════════════════════════════════════════════
    original_df  = st.session_state.df
    anomaly_list = st.session_state.anomaly_list
    dup_list     = st.session_state.duplicate_list
    anomaly_set  = st.session_state.anomaly_set

    # ════════════════════════════════════════════════
    # 模式分流
    # ════════════════════════════════════════════════
    if app_mode == '資料整理':
        _render_data_mode(original_df, anomaly_list, dup_list, anomaly_set)
    else:
        _render_gml_mode_tabs()


# ============================================================
# 資料整理模式
# ============================================================

def _render_data_mode(original_df, anomaly_list, dup_list, anomaly_set):
    # ── 主畫面：畫布 + 編輯 ────────────────────────
    col_canvas, col_edit = st.columns([3, 2])

    with col_canvas:
        fig = build_figure_v3(
            st.session_state.edited_points_df,
            anomaly_set=anomaly_set,
            selected_key=st.session_state.selected_key,
            canvas_height=580,
            uirevision=st.session_state.canvas_revision,
            show_excluded=st.session_state.get('show_excluded', False),
        )
        event = st.plotly_chart(
            fig, use_container_width=True,
            config={'scrollZoom': True, 'displayModeBar': True,
                    'toImageButtonOptions': {'format': 'png', 'filename': 'GML_點位畫布_v0_6',
                                             'height': 1000, 'width': 1600}},
            on_select='rerun', selection_mode='points', key='v3_canvas',
        )
        new_key = extract_point_key(event)
        if new_key and new_key != st.session_state.selected_key:
            st.session_state.selected_key = new_key
            st.session_state.confirm_clear = False
            st.rerun()

        edf    = st.session_state.edited_points_df
        n_mod  = int(edf['is_modified'].sum())
        n_excl = int(edf.apply(_is_excl, axis=1).sum())
        n_anom = sum(1 for item in anomaly_list
                     if str(item.get('原始列號', '')) != '（整份檔案）')
        st.caption(
            f'**{len(edf)}** 個點位　異常 **{n_anom}** 個　'
            f'已修正 **{n_mod}** 個　已排除 **{n_excl}** 個　'
            '｜ **點選畫布上的點位** 即可在右側編輯')

    with col_edit:
        render_edit_panel(st.session_state.edited_points_df, original_df, anomaly_set)

    # ── 下方頁籤 ─────────────────────────────────
    st.divider()
    tabs = st.tabs([
        '📋 修正後點位表',
        '📋 原始解析點位表',
        '⚠️ 異常清單',
        '🔄 重複量測比對',
        '📝 修正紀錄',
        '📊 分類表',
        '🔗 設施屬性合併',
        '📥 下載',
    ])

    # Tab 0：修正後點位表
    with tabs[0]:
        _render_tab_edited()

    # Tab 1：原始解析
    with tabs[1]:
        st.subheader('原始解析點位表（唯讀）')
        st.caption('此表格永遠反映原始解析結果，不受任何人工修正影響。')
        disp = original_df[[
            '_source_file', '_original_row', 'raw_id', 'pipeline_id', 'point_no',
            'raw_category', 'raw_x', 'raw_y', 'raw_z', 'raw_date',
        ]].copy()
        disp.columns = ['來源CSV', '原始列號', '原始識別碼', '管線識別碼', '點號',
                        '類別碼', 'X', 'Y', 'Z', '測量日期']
        st.dataframe(disp, use_container_width=True, hide_index=True)
        st.caption(f'共 {len(disp)} 筆')

    # Tab 2：異常清單
    with tabs[2]:
        st.subheader('異常清單')
        if anomaly_list:
            anom_df = pd.DataFrame(anomaly_list)
            type_counts = anom_df['異常類型'].value_counts()
            cols = st.columns(min(len(type_counts), 4))
            for idx, (atype, cnt) in enumerate(type_counts.items()):
                cols[idx % 4].metric(atype, cnt)
            st.divider()
            st.dataframe(make_anomaly_styler(anom_df), use_container_width=True, hide_index=True)
            st.caption('紅底 = 高嚴重度異常')
        else:
            st.success('✅ 未偵測到任何異常！')

    # Tab 3：重複量測比對
    with tabs[3]:
        st.subheader('重複量測比對表')
        st.caption(f'判定門檻：XY ≤ {st.session_state.xy_tol_used:.2f}m　'
                   f'Z ≤ {st.session_state.z_tol_used:.2f}m')
        if dup_list:
            dup_df = pd.DataFrame(dup_list)
            def _style_dup(row):
                v = str(row.get('判定結果', ''))
                if '差異過大' in v: return ['background-color: #fdecea'] * len(row)
                if '近似重複' in v: return ['background-color: #fef9e7'] * len(row)
                return [''] * len(row)
            st.dataframe(dup_df.style.apply(_style_dup, axis=1),
                         use_container_width=True, hide_index=True)
            st.caption('紅底 = 差異過大　黃底 = 近似重複')
        else:
            st.success('✅ 未偵測到重複量測！')

    # Tab 4：修正紀錄
    with tabs[4]:
        st.subheader('修正紀錄表')
        st.caption('所有「儲存修改」、「快速標記」、「還原此點」、「清除全部修正」操作，'
                   '均自動寫入此表。')
        cur_log = st.session_state.edit_log_df
        if cur_log is not None and not cur_log.empty:
            st.dataframe(cur_log, use_container_width=True, hide_index=True)
            st.caption(f'共 {len(cur_log)} 筆修正紀錄')
        else:
            st.info('目前尚無修正紀錄。在畫布點選點位並執行修改後，紀錄將出現在此。')

    # Tab 5：分類表
    with tabs[5]:
        _render_tab_classification()

    # Tab 6：設施屬性合併
    with tabs[6]:
        _render_tab_attribute_merge()

    # Tab 7：下載
    with tabs[7]:
        _render_tab_download(original_df, anomaly_list, dup_list, anomaly_set)


def _render_tab_edited():
    show_excl = st.session_state.get('show_excluded', False)
    edf_raw   = st.session_state.edited_points_df
    edf_filt  = edf_raw if show_excl else edf_raw[~edf_raw.apply(_is_excl, axis=1)]
    pkeys_t0  = edf_filt['point_key'].tolist()

    _SRC_COLS  = ['_source_file','_original_row','raw_id','edited_raw_id',
                  'edited_category','edited_pipeline_id','edited_point_no',
                  'edited_x','edited_y','edited_z','edited_date',
                  'review_status','problem_type','manual_note','include_in_result','is_modified']
    _DISP_COLS = ['來源CSV','列號','原始識別碼','識別碼','類別碼','管線識別碼','點號',
                  'X','Y','Z','測量日期','判讀狀態','問題類型','備註','納入成果','已修正']

    ed_df = edf_filt[_SRC_COLS].copy().reset_index(drop=True)
    ed_df.columns = _DISP_COLS
    ed_df['已修正'] = ed_df['已修正'].apply(lambda v: '⚙️' if v else '')
    ed_df.insert(0, '刪除', False)

    st.subheader('修正後點位表')
    n_all_t0  = len(edf_raw)
    n_excl_t0 = int(edf_raw.apply(_is_excl, axis=1).sum())
    n_mod_t0  = int(edf_raw['is_modified'].sum())
    st.caption(
        f'顯示 **{len(edf_filt)}** / {n_all_t0} 筆　'
        f'已排除 **{n_excl_t0}** 筆{"（已顯示）" if show_excl else "（已隱藏）"}　'
        f'已修正 **{n_mod_t0}** 筆　｜　勾選「🗑️」列後點擊下方按鈕排除')

    edited_t0 = st.data_editor(
        ed_df,
        column_config={'刪除': st.column_config.CheckboxColumn('🗑️', default=False,
                                                                help='勾選欲排除的列')},
        disabled=[c for c in ed_df.columns if c != '刪除'],
        hide_index=True, use_container_width=True, key='tab0_editor',
    )

    checked_t0 = [i for i, v in enumerate(edited_t0['刪除'].tolist()) if v]
    to_excl_t0 = [pkeys_t0[i] for i in checked_t0 if i < len(pkeys_t0)]

    if st.session_state.get('confirm_excl_tab0') and st.session_state.get('pending_excl_tab0'):
        pending_t0 = st.session_state['pending_excl_tab0']
        st.warning(f'⚠️ 確定排除 **{len(pending_t0)}** 筆？')
        cc1, cc2 = st.columns(2)
        if cc1.button('✅ 確認排除', key='t0_excl_yes', type='primary', use_container_width=True):
            new_edf, new_log = exclude_points(
                st.session_state.edited_points_df, st.session_state.edit_log_df, pending_t0)
            st.session_state.edited_points_df = new_edf
            st.session_state.edit_log_df = new_log
            st.session_state.confirm_excl_tab0 = False
            st.session_state.pending_excl_tab0 = []
            st.toast(f'🗑️ 已排除 {len(pending_t0)} 筆', icon='🗑️')
            st.rerun()
        if cc2.button('❌ 取消', key='t0_excl_no', use_container_width=True):
            st.session_state.confirm_excl_tab0 = False
            st.session_state.pending_excl_tab0 = []
            st.rerun()
    else:
        btn_col, info_col = st.columns([2, 5])
        with btn_col:
            if st.button(f'🗑️ 排除勾選（{len(to_excl_t0)} 筆）',
                         disabled=not to_excl_t0, key='t0_excl_btn', use_container_width=True):
                st.session_state.confirm_excl_tab0 = True
                st.session_state.pending_excl_tab0 = to_excl_t0
                st.rerun()
        with info_col:
            if to_excl_t0:
                st.info(f'已勾選 {len(to_excl_t0)} 筆，點擊按鈕後確認排除。')


def _render_tab_classification():
    st.subheader('分類表')
    st.caption('依類別碼自動判斷設施類型。空白欄位需人工補填。已排除資料預設不顯示。')
    show_excl_cls = st.session_state.get('show_excluded', False)
    edf_cls = st.session_state.edited_points_df.copy()

    def _safe_ftype(c):
        s = str(c).strip() if (c is not None and not (isinstance(c, float) and math.isnan(c))) else ''
        return get_facility_type(s)

    edf_cls_stat = edf_cls if show_excl_cls else edf_cls[~edf_cls.apply(_is_excl, axis=1)]
    edf_cls_stat = edf_cls_stat.copy()
    edf_cls_stat['__ftype__'] = edf_cls_stat['edited_category'].apply(_safe_ftype)
    ftype_counts = edf_cls_stat['__ftype__'].value_counts()

    st.markdown('**各設施類型點位數量：**')
    stat_cols = st.columns(len(FACILITY_TYPES))
    for i, ft in enumerate(FACILITY_TYPES):
        stat_cols[i].metric(ft, int(ftype_counts.get(ft, 0)))

    st.divider()

    sel_idx = st.radio('選擇設施分類', range(len(FACILITY_TYPES)),
        format_func=lambda i: FACILITY_TYPES[i], horizontal=True, key='cls_sel_type')
    sel_type = FACILITY_TYPES[sel_idx]

    cls_df    = build_classification_df(edf_cls, sel_type, include_excluded=show_excl_cls)
    cls_pkeys = build_cls_pkey_list(edf_cls, sel_type, include_excluded=show_excl_cls)
    n_rows    = len(cls_df)

    if n_rows == 0:
        st.info(f'目前資料中沒有「{sel_type}」類型的點位。')
    else:
        unit = '條' if sel_type == '管線' else '個'
        st.caption(f'{sel_type}｜共 {n_rows} {unit}｜欄位數：{len(cls_df.columns)}')

        cls_ed = cls_df.copy().reset_index(drop=True)
        # 管線型：將 第N點X/Y/Z/d rename 為 E/N/Z/d（加序號避免重複）
        if sel_type == '管線':
            rename_map = {}
            for c in cls_ed.columns:
                for suffix, repl in [('X', 'E'), ('Y', 'N'), ('Z', 'Z'), ('d', 'd')]:
                    if c.startswith('第') and c.endswith(f'點{suffix}'):
                        n = c[1:c.index('點')]
                        rename_map[c] = f'{repl}({n})'
                        break
            cls_ed = cls_ed.rename(columns=rename_map)
        cls_ed.insert(0, '刪除', False)

        edited_cls = st.data_editor(
            cls_ed,
            column_config={'刪除': st.column_config.CheckboxColumn('🗑️', default=False,
                help='勾選後點擊「排除勾選」')},
            disabled=[c for c in cls_ed.columns if c != '刪除'],
            hide_index=True, use_container_width=True, key=f'cls_editor_{sel_type}',
        )

        checked_cls = [i for i, v in enumerate(edited_cls['刪除'].tolist()) if v]
        to_excl_cls = []
        for i in checked_cls:
            if i >= len(cls_pkeys): continue
            entry = cls_pkeys[i]
            if isinstance(entry, list): to_excl_cls.extend(entry)
            else: to_excl_cls.append(entry)

        if st.session_state.get('confirm_excl_cls') and st.session_state.get('pending_excl_cls'):
            pending_cls = st.session_state['pending_excl_cls']
            st.warning(f'⚠️ 確定排除 **{len(pending_cls)}** 個點位？')
            cc1, cc2 = st.columns(2)
            if cc1.button('✅ 確認排除', key='cls_excl_yes', type='primary', use_container_width=True):
                new_edf, new_log = exclude_points(
                    st.session_state.edited_points_df, st.session_state.edit_log_df, pending_cls)
                st.session_state.edited_points_df = new_edf
                st.session_state.edit_log_df = new_log
                st.session_state.confirm_excl_cls = False
                st.session_state.pending_excl_cls = []
                st.toast(f'🗑️ 已排除 {len(pending_cls)} 個點位', icon='🗑️')
                st.rerun()
            if cc2.button('❌ 取消', key='cls_excl_no', use_container_width=True):
                st.session_state.confirm_excl_cls = False
                st.session_state.pending_excl_cls = []
                st.rerun()
        else:
            btn_c, info_c = st.columns([2, 5])
            with btn_c:
                if st.button(f'🗑️ 排除勾選（{len(to_excl_cls)} 點）',
                             disabled=not to_excl_cls, key='cls_excl_btn', use_container_width=True):
                    st.session_state.confirm_excl_cls = True
                    st.session_state.pending_excl_cls = to_excl_cls
                    st.rerun()
            with info_c:
                if to_excl_cls:
                    st.info(f'管線排除將影響該管線的所有 {len(to_excl_cls)} 個點位。')

    st.caption('d = 埋設深度（公尺）。管線表每條管線佔一列，點位座標以「第N點X/Y/Z/d」橫向展開。')


def _render_tab_attribute_merge():
    st.subheader('設施屬性合併')
    st.caption(
        '上傳人工填寫的「分類表_定義版」Excel，系統將自動與點位分類表合併屬性，'
        '並產生定義版、代碼版、代碼定義對照版三種成果。')

    code_maps = load_code_mappings(CONFIG_PATH)

    uploaded_def = st.file_uploader(
        '上傳分類表_定義版 Excel（.xlsx）',
        type=['xlsx'],
        key='attr_merge_upload',
        help='工作表名稱可為：管線 / 分類表_管線 / 分類表_管線_定義版 等，系統會自動辨識。',
    )

    if uploaded_def is None:
        st.info('請上傳人工填寫的分類表_定義版 Excel，系統將與目前點位分類表合併。')
        _render_attr_merge_flow_chart()
        return

    def_sheets, err = load_definition_excel(uploaded_def)
    if err:
        st.error(err)
        return
    if not def_sheets:
        st.warning('未在 Excel 中找到可辨識的設施分類工作表。')
        return

    st.success(f'已辨識 {len(def_sheets)} 個設施類型：{", ".join(def_sheets.keys())}')

    edf = st.session_state.edited_points_df
    all_merged = {}
    all_unmatched = {}
    all_unfilled = {}
    all_code_errors = []
    total_merged = 0

    for ft, def_df in def_sheets.items():
        cls_df = build_classification_df(edf, ft, include_excluded=False)
        if cls_df.empty:
            continue
        merged, unmatched, unfilled = merge_definition_attributes(cls_df, def_df, ft)
        all_merged[ft] = merged
        if not unmatched.empty:
            all_unmatched[ft] = unmatched
        if not unfilled.empty:
            all_unfilled[ft] = unfilled
        total_merged += len(merged)

    if not all_merged:
        st.warning('沒有任何設施類型成功合併。請確認定義版工作表與目前資料的設施類型是否一致。')
        return

    st.session_state['attr_merged'] = all_merged

    st.divider()
    st.markdown('#### 合併結果')

    mc1, mc2, mc3 = st.columns(3)
    mc1.metric('成功合併', f'{total_merged} 筆')
    mc2.metric('未匹配屬性', f'{sum(len(v) for v in all_unmatched.values())} 筆')
    mc3.metric('屬性未填', f'{sum(len(v) for v in all_unfilled.values())} 筆')

    for ft, merged_df in all_merged.items():
        with st.expander(f'{ft}（{len(merged_df)} 筆）', expanded=False):
            st.dataframe(merged_df, use_container_width=True, hide_index=True)

    if all_unmatched:
        with st.expander(f'⚠️ 未匹配屬性資料（定義版有，點位表無）', expanded=False):
            for ft, udf in all_unmatched.items():
                st.caption(f'**{ft}**：{len(udf)} 筆')
                st.dataframe(udf, use_container_width=True, hide_index=True)

    if all_unfilled:
        with st.expander(f'⚠️ 屬性未填資料（點位表有，定義版無）', expanded=False):
            for ft, udf in all_unfilled.items():
                st.caption(f'**{ft}**：{len(udf)} 筆')
                st.dataframe(udf, use_container_width=True, hide_index=True)

    # ── 產生三版本 ─────────────────────────────────
    st.divider()
    st.markdown('#### 下載三版本成果')

    all_code_dfs = {}
    all_convert_errors = []
    for ft, mdf in all_merged.items():
        code_df, errs = convert_definition_to_code(mdf, code_maps)
        all_code_dfs[ft] = code_df
        for e in errs:
            e['facility'] = ft
        all_convert_errors.extend(errs)

    if all_convert_errors:
        st.warning(f'代碼轉換異常：{len(all_convert_errors)} 筆')
        with st.expander('代碼轉換異常明細', expanded=False):
            st.dataframe(pd.DataFrame(all_convert_errors), use_container_width=True, hide_index=True)

    dc1, dc2, dc3 = st.columns(3)

    # 定義版
    with dc1:
        def_buf = _export_merged_excel(all_merged, '定義版')
        st.download_button('📥 下載分類表_定義版',
            data=def_buf, file_name='分類表_定義版.xlsx', mime=_MIME_XLSX,
            use_container_width=True, key='dl_attr_def')

    # 代碼版
    with dc2:
        code_buf = _export_merged_excel(all_code_dfs, '代碼版')
        st.download_button('📥 下載分類表_代碼版',
            data=code_buf, file_name='分類表_代碼版.xlsx', mime=_MIME_XLSX,
            use_container_width=True, key='dl_attr_code')

    # 對照版
    with dc3:
        ref_dfs = {}
        for ft, cdf in all_code_dfs.items():
            ref_dfs[ft] = build_code_definition_table(cdf, code_maps)
        ref_buf = _export_merged_excel(ref_dfs, '對照版')
        st.download_button('📥 下載分類表_對照版',
            data=ref_buf, file_name='分類表_代碼定義對照版.xlsx', mime=_MIME_XLSX,
            use_container_width=True, key='dl_attr_ref')

    st.session_state['attr_code_dfs'] = all_code_dfs


def _export_merged_excel(dfs_dict, version_label):
    """將 {設施類型: DataFrame} 輸出為多工作表 Excel BytesIO。"""
    wb = openpyxl.Workbook()
    first = True
    for ft, df in dfs_dict.items():
        if first:
            ws = wb.active
            first = False
        else:
            ws = wb.create_sheet()
        ws.title = f'{ft}_{version_label}'
        if df is not None and not df.empty:
            headers = list(df.columns)
            for c, h in enumerate(headers, 1):
                ws.cell(row=1, column=c, value=h)
            for r, (_, row) in enumerate(df.iterrows(), 2):
                for c, h in enumerate(headers, 1):
                    v = row[h]
                    if v is None or (isinstance(v, float) and math.isnan(v)):
                        v = ''
                    ws.cell(row=r, column=c, value=v)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _render_attr_merge_flow_chart():
    st.markdown("""
---
**正式流程：**

1. 工具完成點位整理 → 產生分類表點位資料
2. 下載分類表（下載頁籤），人工填寫屬性欄位 → 另存為「分類表_定義版」
3. 上傳「分類表_定義版」回本頁
4. 系統自動合併屬性 + 產生定義版/代碼版/對照版
5. GML 輸出使用代碼版

**人工建議填定義內容（如「使用」、「地下」、「mm」），不建議直接填代碼。**
工具會自動將定義文字轉為標準代碼。
    """)


def _build_versioned_cls_bytes(edited_df, facility_type, version, code_maps):
    """
    產生指定版本的分類表 Excel bytes。
    version: '定義版' / '代碼版' / '對照版'
    """
    from src.v3_helpers import pipeline_export_rename, _write_sheet, _CLS_SHEET_FILLS

    types_to_export = FACILITY_TYPES if facility_type == '全部設施' else [facility_type]
    wb = openpyxl.Workbook()
    first = True

    for idx, ft in enumerate(types_to_export):
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = f'{ft}_{version}'
        cls_df = build_classification_df(edited_df, ft, include_excluded=False)
        if cls_df.empty:
            continue

        if version == '定義版':
            out_df = convert_code_to_definition(cls_df, code_maps)
        elif version == '代碼版':
            out_df, _ = convert_definition_to_code(cls_df, code_maps)
        else:  # 對照版
            code_df, _ = convert_definition_to_code(cls_df, code_maps)
            out_df = build_code_definition_table(code_df, code_maps)

        fill = _CLS_SHEET_FILLS[FACILITY_TYPES.index(ft)] if ft in FACILITY_TYPES else 'blue'
        if ft == '管線':
            headers, rows = pipeline_export_rename(out_df)
            _write_sheet(ws, headers, rows, fill)
        else:
            _write_sheet(ws, list(out_df.columns), out_df.to_dict('records'), fill)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _render_tab_download(original_df, anomaly_list, dup_list, anomaly_set):
    st.subheader('下載成果檔案')

    # 1. 整合成果 Excel
    _dl_code_maps = load_code_mappings(CONFIG_PATH)
    st.markdown('#### 📊 整合成果 Excel')
    v_bytes = export_v3_bytes(
        original_df, st.session_state.edited_points_df,
        st.session_state.edit_log_df, anomaly_list, dup_list,
        code_mappings=_dl_code_maps)
    n_mod_dl = int(st.session_state.edited_points_df['is_modified'].sum())
    n_log_dl = len(st.session_state.edit_log_df) if st.session_state.edit_log_df is not None else 0
    st.download_button(
        f'📥 GML處理成果_v0_6.xlsx　（含 {n_mod_dl} 筆修正、{n_log_dl} 筆紀錄）',
        data=v_bytes, file_name='GML處理成果_v0_6.xlsx', mime=_MIME_XLSX)
    n_excl_dl = int(st.session_state.edited_points_df.apply(_is_excl, axis=1).sum())
    st.caption(
        f'修正後點位表 / 人工判讀紀錄表 / 分類表：僅含納入成果的點位。'
        f'已排除 {n_excl_dl} 筆資料另存於「已排除資料」工作表。　'
        '工作表（16）：原始解析 / 修正後成果 / 人工判讀 / 修正紀錄 / 異常清單 / 重複量測 '
        '/ 分類表×9 / 已排除資料')

    st.divider()

    # 2. 分類表 Excel 下載（三版本）
    st.markdown('#### 📂 分類表 Excel 下載')
    st.caption('每種版本均含全部設施與各設施個別下載。定義版供 BIM/人工檢查，代碼版供 GML/廠商，對照版供查核。')
    _cls_dl_edf = st.session_state.edited_points_df
    _code_maps = load_code_mappings(CONFIG_PATH)

    ver_tabs = st.tabs(['📄 定義版', '🔢 代碼版', '📋 對照版'])

    for ver_idx, (ver_tab, ver_label, ver_suffix) in enumerate(zip(
        ver_tabs,
        ['定義版', '代碼版', '對照版'],
        ['定義版', '代碼版', '對照版'],
    )):
        with ver_tab:
            st.download_button(
                f'📥 下載全部分類_{ver_label} Excel',
                data=_build_versioned_cls_bytes(_cls_dl_edf, '全部設施', ver_suffix, _code_maps),
                file_name=f'分類表_全部設施_{ver_suffix}.xlsx', mime=_MIME_XLSX,
                use_container_width=True, key=f'dl_cls_all_{ver_suffix}')
            _cls_rows = [FACILITY_TYPES[i:i+3] for i in range(0, len(FACILITY_TYPES), 3)]
            for _row_types in _cls_rows:
                _btn_cols = st.columns(3)
                for _col, _ft in zip(_btn_cols, _row_types):
                    with _col:
                        st.download_button(
                            f'📥 {_ft}_{ver_label}',
                            data=_build_versioned_cls_bytes(_cls_dl_edf, _ft, ver_suffix, _code_maps),
                            file_name=f'分類表_{_ft}_{ver_suffix}.xlsx', mime=_MIME_XLSX,
                            use_container_width=True, key=f'dl_cls_{_ft}_{ver_suffix}')

    st.divider()

    # 3. 分析成果個別檔 + 4. HTML + 5. 範本
    col1, col2, col3 = st.columns(3)

    with col1:
        st.markdown('##### 分析成果（個別）')
        st.download_button('📥 解析後點位表.xlsx',
            data=_export_to_bytes(export_parsed_points, original_df),
            file_name='解析後點位表.xlsx', mime=_MIME_XLSX, use_container_width=True)
        st.download_button('📥 異常清單.xlsx',
            data=_export_to_bytes(export_anomalies, anomaly_list),
            file_name='異常清單.xlsx', mime=_MIME_XLSX, use_container_width=True)
        st.download_button('📥 重複量測比對表.xlsx',
            data=_export_to_bytes(export_duplicates, dup_list),
            file_name='重複量測比對表.xlsx', mime=_MIME_XLSX, use_container_width=True)

    with col2:
        st.markdown('##### 點位畫布 HTML')
        fig_dl = build_figure_v3(
            st.session_state.edited_points_df, anomaly_set=anomaly_set,
            selected_key=None, canvas_height=800, show_excluded=False)
        html_str = pio.to_html(fig_dl, include_plotlyjs=True, full_html=True,
                               config={'scrollZoom': True})
        st.download_button('📥 點位畫布_v0_6.html（修正後）',
            data=html_str.encode('utf-8'), file_name='點位畫布_v0_6.html',
            mime='text/html', use_container_width=True)
        st.caption('HTML 可離線用瀏覽器開啟，反映最新修正結果。')

        st.markdown('##### 空白範本')
        st.download_button('📥 人工判讀紀錄範本.xlsx',
            data=_export_to_bytes(export_review_template, original_df, anomaly_set),
            file_name='人工判讀紀錄範本.xlsx', mime=_MIME_XLSX, use_container_width=True)
        st.download_button('📥 修正紀錄範本.xlsx（空白）',
            data=_export_to_bytes(export_correction_template),
            file_name='修正紀錄範本.xlsx', mime=_MIME_XLSX, use_container_width=True)

    with col3:
        st.markdown('##### 修正紀錄（目前工作階段）')
        log_buf = export_v3_bytes(
            original_df, st.session_state.edited_points_df,
            st.session_state.edit_log_df, anomaly_list, dup_list,
            code_mappings=_dl_code_maps)
        st.download_button('📥 修正後成果_完整版.xlsx',
            data=log_buf, file_name='修正後成果_完整版.xlsx', mime=_MIME_XLSX,
            use_container_width=True)
        st.caption(f'目前工作階段共有 {n_mod_dl} 筆點位已修正，{n_log_dl} 筆修正操作紀錄。\n'
                   '確認完成後再下載，確保資料完整。')

    st.divider()
    st.info('**提醒：** 下載檔案反映「下載當下」的最新修正狀態。'
            '重新上傳 CSV 或重新整理頁面會清除所有人工修正，請先下載備份。', icon='💡')


# ============================================================
# GML 輸出模式
# ============================================================

def _render_gml_mode_tabs():
    gml_tabs = st.tabs([
        '🔧 GML 產製',
        '👁️ GML 預覽',
        '📥 GML 下載',
        '📋 GML 檢核訊息',
    ])

    gml_source    = st.session_state.get('gml_source_sb', '使用目前修正後資料')
    gml_target    = list(TARGET_OPTIONS.keys())[st.session_state.get('gml_target_sb', 0)]
    gml_facility  = st.session_state.get('gml_facility_sb', '管線')
    gml_elevation = list(ELEVATION_OPTIONS.keys())[st.session_state.get('gml_elev_sb', 0)]

    target_suffix_map = {'Taoyuan': '桃園市', 'NLMA': '國土署', 'STEC': '南科'}

    # ── Tab 0：GML 產製 ─────────────────────────────
    with gml_tabs[0]:
        st.subheader('GML 產製')

        st.markdown(f"""
**目前設定：**
- 資料來源：{gml_source}
- 目標規範：{TARGET_OPTIONS.get(gml_target, gml_target)}
- 設施類型：{gml_facility}
- 高程模式：{ELEVATION_OPTIONS.get(gml_elevation, gml_elevation)}
        """)

        st.divider()

        gml_cls_df = None

        if gml_source == '上傳 GML 專用寬表 CSV':
            gml_csv_file = st.file_uploader('上傳 GML 專用寬表 CSV', type=['csv'], key='gml_csv_upload_v0_6')
            if gml_csv_file is not None:
                try:
                    gml_cls_df = pd.read_csv(gml_csv_file, dtype=str, keep_default_na=False)
                    st.success(f'已載入 {len(gml_cls_df)} 筆資料')
                except Exception as e:
                    st.error(f'CSV 讀取失敗：{e}')
        elif st.session_state.processed:
            # 優先使用代碼版（若已完成屬性合併）
            attr_code_dfs = st.session_state.get('attr_code_dfs', {})
            if gml_facility in attr_code_dfs and not attr_code_dfs[gml_facility].empty:
                gml_cls_df = attr_code_dfs[gml_facility]
                st.success(f'使用代碼版資料，{gml_facility}共 {len(gml_cls_df)} 筆（已完成屬性合併與代碼轉換）')
            else:
                gml_cls_df = build_classification_df(
                    st.session_state.edited_points_df, gml_facility, include_excluded=False)
                st.info(f'使用修正後資料（尚未合併定義版屬性），{gml_facility}分類表共 {len(gml_cls_df)} 筆')
            n_excl_gml = int(st.session_state.edited_points_df.apply(_is_excl, axis=1).sum())
            if n_excl_gml:
                st.caption(f'已排除 {n_excl_gml} 筆不含在內')

            if gml_cls_df is not None and not gml_cls_df.empty:
                csv_bytes = export_gml_wide_csv_bytes(gml_cls_df, gml_facility)
                st.download_button(
                    f'📥 下載 GML 專用寬表 CSV（{gml_facility}）',
                    data=csv_bytes, file_name=f'GML寬表_{gml_facility}.csv',
                    mime='text/csv', key='dl_gml_wide_csv_v0_6')
        else:
            st.warning('請先完成資料整理，或切換資料來源為「上傳 GML 專用寬表 CSV」。')

        st.divider()

        if not is_implemented(gml_target, gml_facility):
            st.warning(f'⚠️ {TARGET_OPTIONS.get(gml_target, gml_target)} + {gml_facility}：'
                       f'此規範或設施類型尚未實作。\n目前支援：桃園市 + 全部設施類型。')
        elif gml_cls_df is not None and not gml_cls_df.empty:
            if st.button('🚀 產生 GML', type='primary', key='btn_gen_gml_v0_6'):
                with st.spinner('正在產製 GML...'):
                    gml_text, gml_ok, gml_errors = generate_gml(
                        gml_cls_df, gml_target, gml_facility, gml_elevation)

                st.session_state.gml_result_text = gml_text
                st.session_state.gml_result_ok = gml_ok
                st.session_state.gml_result_errors = gml_errors
                st.session_state.gml_result_facility = gml_facility
                st.session_state.gml_result_target = gml_target

                if gml_ok > 0:
                    st.success(f'GML 產製完成：成功 {gml_ok} 筆設施')
                else:
                    st.error('GML 產製失敗：0 筆設施成功')
        elif gml_cls_df is not None and gml_cls_df.empty:
            st.info(f'目前資料中沒有「{gml_facility}」類型的點位。')

    # ── Tab 1：GML 預覽 ─────────────────────────────
    with gml_tabs[1]:
        st.subheader('GML 預覽')
        gml_text = st.session_state.get('gml_result_text', '')
        if gml_text:
            st.caption(f'設施類型：{st.session_state.gml_result_facility}　'
                       f'目標規範：{TARGET_OPTIONS.get(st.session_state.gml_result_target, "")}　'
                       f'成功 {st.session_state.gml_result_ok} 筆')
            preview_len = st.slider('預覽字元數', 1000, min(50000, len(gml_text)),
                                    value=min(5000, len(gml_text)), step=1000, key='gml_preview_len')
            st.code(gml_text[:preview_len], language='xml')
            if len(gml_text) > preview_len:
                st.caption(f'（僅顯示前 {preview_len} 字元，完整檔案共 {len(gml_text)} 字元）')
        else:
            st.info('請先在「GML 產製」頁籤產生 GML。')

    # ── Tab 2：GML 下載 ─────────────────────────────
    with gml_tabs[2]:
        st.subheader('GML 下載')
        gml_text = st.session_state.get('gml_result_text', '')
        if gml_text:
            target_label = target_suffix_map.get(st.session_state.gml_result_target,
                                                  st.session_state.gml_result_target)
            facility_label = st.session_state.gml_result_facility
            gml_filename = f'GML_{target_label}_{facility_label}.gml'

            st.download_button(
                f'📥 下載 {gml_filename}',
                data=gml_text.encode('utf-8'),
                file_name=gml_filename,
                mime='application/gml+xml',
                key='dl_gml_file_v0_6',
                use_container_width=True,
            )
            st.caption(f'檔案大小：{len(gml_text):,} 字元　'
                       f'成功 {st.session_state.gml_result_ok} 筆設施')
        else:
            st.info('請先在「GML 產製」頁籤產生 GML。')

    # ── Tab 3：GML 檢核訊息 ─────────────────────────
    with gml_tabs[3]:
        st.subheader('GML 檢核訊息')
        gml_ok = st.session_state.get('gml_result_ok', 0)
        gml_errors = st.session_state.get('gml_result_errors', [])
        gml_text = st.session_state.get('gml_result_text', '')

        if not gml_text and not gml_errors:
            st.info('請先在「GML 產製」頁籤產生 GML。')
        else:
            c1, c2, c3 = st.columns(3)
            c1.metric('成功產製', f'{gml_ok} 筆')

            if st.session_state.processed:
                n_excl_check = int(st.session_state.edited_points_df.apply(_is_excl, axis=1).sum())
                c2.metric('排除資料', f'{n_excl_check} 筆')
            c3.metric('產製異常', f'{len(gml_errors)} 筆')

            if gml_errors:
                st.divider()
                st.markdown('#### 異常明細')
                err_df = pd.DataFrame(gml_errors)
                st.dataframe(err_df, use_container_width=True, hide_index=True)

                err_types = {}
                for e in gml_errors:
                    reason = e.get('error', '未知')
                    if '類別碼' in reason: err_types['類別碼缺漏或無效'] = err_types.get('類別碼缺漏或無效', 0) + 1
                    elif '座標' in reason: err_types['座標缺漏'] = err_types.get('座標缺漏', 0) + 1
                    else: err_types['其他'] = err_types.get('其他', 0) + 1

                st.markdown('**異常類型統計：**')
                for etype, cnt in err_types.items():
                    st.caption(f'• {etype}：{cnt} 筆')
            elif gml_ok > 0:
                st.success('✅ 所有設施均成功產製，無異常。')


if __name__ == '__main__':
    main()
